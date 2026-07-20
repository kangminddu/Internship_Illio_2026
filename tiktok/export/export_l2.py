# tiktok/export/export_l2.py
import os
import warnings

warnings.filterwarnings("ignore", category=UserWarning)

import pandas as pd
import pymysql
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from tiktok.config import DB

# 출력 위치
OUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "output",
)
os.makedirs(OUT_DIR, exist_ok=True)

sql = """
SELECT
  ROW_NUMBER() OVER(
    ORDER BY cs.view_count DESC
  ) AS '번호',
  cr.nickname AS '크리에이터',
  chs.follower_count AS '팔로워',
  DATE_FORMAT(ct.published_at, '%Y-%m-%d %H:%i') AS '게시일',
  DATEDIFF(CURDATE(), ct.published_at) AS '업로드경과일',
  cs.view_count AS '조회수',
  cs.like_count AS '좋아요',
  cs.comment_count AS '댓글수',
  ROUND(cs.like_count / NULLIF(cs.view_count,0) * 100, 2) AS '좋아요율(%)',
  ROUND(cs.comment_count / NULLIF(cs.view_count,0) * 100, 2) AS '댓글율(%)',
  ct.duration_sec AS '길이(초)',
  ct.caption_text AS '캡션',
  CONCAT(ch.channel_url_normalized, '/video/', ct.external_id) AS '영상URL'
FROM creators cr
JOIN channels ch
  ON cr.creator_id = ch.creator_id
JOIN contents ct
  ON ch.channel_id = ct.channel_id
LEFT JOIN (
  SELECT cs1.*
  FROM content_snapshots cs1
  JOIN (
    SELECT content_id, MAX(captured_at) latest
    FROM content_snapshots
    GROUP BY content_id
  ) t
    ON cs1.content_id = t.content_id
   AND cs1.captured_at = t.latest
) cs
  ON ct.content_id = cs.content_id
LEFT JOIN (
  SELECT chs1.channel_id, chs1.follower_count
  FROM channel_snapshots chs1
  JOIN (
    SELECT channel_id, MAX(captured_at) latest
    FROM channel_snapshots
    GROUP BY channel_id
  ) t2
    ON chs1.channel_id = t2.channel_id
   AND chs1.captured_at = t2.latest
) chs
  ON ch.channel_id = chs.channel_id
WHERE ch.platform='tiktok'
ORDER BY cs.view_count DESC;
"""


def sanitize_formula(val):
    if isinstance(val, str) and val and val[0] in ("=", "+", "-", "@"):
        return "'" + val
    return val


# ---------------- DB → DataFrame ----------------
conn = pymysql.connect(**DB)
df = pd.read_sql(sql, conn)
conn.close()

for col in ["크리에이터", "캡션"]:
    if col in df.columns:
        df[col] = df[col].apply(sanitize_formula)

# ---------------- 영상길이(mm:ss) ----------------
df["영상길이"] = (
    pd.to_timedelta(df["길이(초)"], unit="s")
    .astype(str)
    .str.replace("0 days ", "", regex=False)
)

df = df.drop(columns=["길이(초)"])

cols = list(df.columns)
cols.remove("영상길이")

idx = cols.index("댓글율(%)") + 1
cols.insert(idx, "영상길이")

df = df[cols]

# ---------------- Excel 저장 ----------------
filename = os.path.join(OUT_DIR, "L2_틱톡_영상정보.xlsx")
df.to_excel(filename, index=False, sheet_name="L2")

# =====================================================
# Styling
# =====================================================

wb = load_workbook(filename)
ws = wb["L2"]

header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF")

center = Alignment(horizontal="center", vertical="center")

thin = Side(border_style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 헤더
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

# 컬럼명 → 열 문자
col_letter = {}
for cell in ws[1]:
    col_letter[cell.value] = cell.column_letter

# 가운데 정렬
center_cols = {
    "번호",
    "팔로워",
    "게시일",
    "업로드경과일",
    "조회수",
    "좋아요",
    "댓글수",
    "좋아요율(%)",
    "댓글율(%)",
    "영상길이",
}

center_letters = {
    col_letter[c]
    for c in center_cols
    if c in col_letter
}

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        if cell.column_letter in center_letters:
            cell.alignment = center

# 숫자 콤마
for name in ["팔로워", "조회수", "좋아요", "댓글수"]:
    if name in col_letter:
        for cell in ws[col_letter[name]][1:]:
            cell.number_format = "#,##0"

# 열 너비
width_map = {
    "번호": 8,
    "크리에이터": 20,
    "팔로워": 12,
    "게시일": 18,
    "업로드경과일": 12,
    "조회수": 12,
    "좋아요": 11,
    "댓글수": 10,
    "좋아요율(%)": 11,
    "댓글율(%)": 11,
    "영상길이": 12,
    "캡션": 55,
    "영상URL": 55,
}

for name, width in width_map.items():
    if name in col_letter:
        ws.column_dimensions[col_letter[name]].width = width

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(filename)

print(f"완료 : {filename}")
print(f"영상 {len(df)}개")