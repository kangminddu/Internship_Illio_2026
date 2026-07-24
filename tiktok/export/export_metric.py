# tiktok/export/export_metric.py
"""
TikTok 파생지표 export

calc_metric.py (L2 지표) + calc_l3_metric.py (L3 지표) 가 채운
channel_metrics 를 그대로 엑셀로 뽑는다. 계산은 하지 않는다.

⚠️ 인스타/유튜브와 달리 TikTok 은 콘텐츠 유형이 하나뿐이라
   롱폼/쇼츠·광고/일반 분리 컬럼이 없다.
⚠️ Loyalty Score 는 (평균댓글*10 + 평균좋아요) / 팔로워 로,
   *100 을 하지 않아 값이 작다. 유튜브/인스타와 직접 비교 불가.
"""

import os
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

import pandas as pd
import pymysql

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from tiktok import config

DB = config.DB
EXPORT_DIR = config.EXPORT_DIR


# =====================================
# SQL — channel_metrics 값 그대로
# =====================================
sql = """
SELECT
    ROW_NUMBER() OVER (ORDER BY cr.nickname) AS 번호,
    cr.nickname AS 크리에이터,
    COALESCE(cr.agency,'-') AS 소속,
    ch.channel_name AS 채널명,
    ch.channel_activity_status AS 활동상태,

    MAX(chs.follower_count) AS follower_count,

    -- 표본 (지표 해석 근거)
    MAX(m.sample_content_count) AS sample_cnt,
    MAX(m.aggregation_method)   AS agg_method,
    MAX(m.videos_3m)            AS videos_3m,
    MAX(m.upload_frequency_weekly) AS upload_freq,

    -- 3개월 대표값
    MAX(m.avg_view_3m)    AS avg_view_3m,
    MAX(m.avg_like_3m)    AS avg_like_3m,
    MAX(m.avg_comment_3m) AS avg_comment_3m,

    -- 비율 지표
    MAX(m.view_per_follower_ratio) AS vpf,
    MAX(m.like_view_ratio)         AS like_ratio,
    MAX(m.comment_view_ratio)      AS comment_ratio,
    MAX(m.engagement_rate)         AS er,
    MAX(m.loyalty_score)           AS loyalty,

    -- L3 댓글 지표 (calc_l3_metric.py 가 UPDATE 로 얹음)
    MAX(m.commenter_overlap_rate)  AS commenter_overlap_rate,
    MAX(m.regular_commenter_count) AS regular_commenter_count,
    MAX(m.avg_comment_length)      AS avg_comment_length,

    MAX(m.calculated_at) AS calculated_at

FROM creators cr

JOIN channels ch
    ON cr.creator_id = ch.creator_id

-- 채널별 최신 스냅샷(팔로워)
LEFT JOIN channel_snapshots chs
    ON ch.channel_id = chs.channel_id
   AND chs.captured_at = (
        SELECT MAX(captured_at)
        FROM channel_snapshots
        WHERE channel_id = ch.channel_id
   )

LEFT JOIN channel_metrics m
    ON ch.channel_id = m.channel_id

WHERE ch.platform='tiktok'

GROUP BY ch.channel_id

ORDER BY cr.nickname;
"""


# =====================================
# SQL → DataFrame
# =====================================
conn = pymysql.connect(**DB)
try:
    df = pd.read_sql(sql, conn)
finally:
    conn.close()

print(f"조회 {len(df):,}행")


# =====================================
# 컬럼 매핑 (계산 없이 값 그대로)
# =====================================
df["팔로워"] = df["follower_count"]
df["활동상태"] = df["활동상태"]

df["표본수"] = df["sample_cnt"]
df["최근3개월영상"] = df["videos_3m"]
df["업로드빈도(주)"] = df["upload_freq"]

df["평균조회수"] = df["avg_view_3m"]
df["평균좋아요"] = df["avg_like_3m"]
df["평균댓글"] = df["avg_comment_3m"]

df["조회수/팔로워(%)"] = df["vpf"]
df["좋아요율(%)"] = df["like_ratio"]
df["댓글율(%)"] = df["comment_ratio"]
df["공개참여율(ER)"] = df["er"]
df["Loyalty Score"] = df["loyalty"]

df["댓글작성자중복률(%)"] = df["commenter_overlap_rate"]
df["고정댓글러"] = df["regular_commenter_count"]
df["평균댓글길이"] = df["avg_comment_length"]

df["집계기준"] = df["agg_method"]
df["계산일시"] = df["calculated_at"]


# =====================================
# 최종 컬럼 순서
# =====================================
FINAL_COLS = [
    "번호",
    "크리에이터",
    "소속",
    "채널명",
    "활동상태",
    "팔로워",

    "표본수",
    "최근3개월영상",
    "업로드빈도(주)",

    "평균조회수",
    "평균좋아요",
    "평균댓글",

    "조회수/팔로워(%)",
    "좋아요율(%)",
    "댓글율(%)",
    "공개참여율(ER)",
    "Loyalty Score",

    "댓글작성자중복률(%)",
    "고정댓글러",
    "평균댓글길이",

    "집계기준",
    "계산일시",
]


# =====================================
# NaN → N/A
#   (지표 미계산 채널을 빈칸이 아니라 명시적으로 표시)
# =====================================
for col in FINAL_COLS:
    df[col] = df[col].apply(lambda x: "N/A" if pd.isna(x) else x)

df = df[FINAL_COLS]


# =====================================
# Excel 저장
# =====================================
os.makedirs(EXPORT_DIR, exist_ok=True)
filename = os.path.join(EXPORT_DIR, "틱톡_파생지표.xlsx")
df.to_excel(filename, index=False, sheet_name="Metrics")

wb = load_workbook(filename)
ws = wb["Metrics"]


# =====================================
# 스타일
# =====================================
header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF")
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
thin = Side(border_style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

# 크리에이터/소속/채널명만 왼쪽 정렬
LEFT_HEADERS = {"크리에이터", "소속", "채널명"}
left_cols = {
    idx for idx, cell in enumerate(ws[1], start=1)
    if cell.value in LEFT_HEADERS
}

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = left if cell.column in left_cols else center


# =====================================
# 헤더명 → 열 문자 (순서 바뀌어도 안 깨짐)
# =====================================
header_to_col = {}
for idx, cell in enumerate(ws[1], start=1):
    header_to_col[cell.value] = get_column_letter(idx)


# =====================================
# 숫자 서식
# =====================================
fmt_map = {
    "팔로워": "#,##0",

    "표본수": "#,##0",
    "최근3개월영상": "#,##0",
    "업로드빈도(주)": "0.00",

    "평균조회수": "#,##0",
    "평균좋아요": "#,##0",
    "평균댓글": "#,##0.0",

    "조회수/팔로워(%)": "0.00",
    "좋아요율(%)": "0.00",
    "댓글율(%)": "0.00",
    "공개참여율(ER)": "0.00",
    # ⚠️ Loyalty 는 팔로워로 나눈 소수라 자릿수를 넉넉히
    "Loyalty Score": "0.0000",

    "댓글작성자중복률(%)": "0.00",
    "고정댓글러": "#,##0",
    "평균댓글길이": "0.0",

    "계산일시": "yyyy-mm-dd hh:mm",
}

for header, fmt in fmt_map.items():
    col_letter = header_to_col.get(header)
    if not col_letter:
        continue
    for cell in ws[col_letter][1:]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = fmt
        elif header == "계산일시" and cell.value not in (None, "N/A"):
            cell.number_format = fmt


# =====================================
# 열 너비
# =====================================
width_map = {
    "번호": 8,
    "크리에이터": 22,
    "소속": 18,
    "채널명": 22,
    "활동상태": 12,
    "팔로워": 14,

    "표본수": 10,
    "최근3개월영상": 14,
    "업로드빈도(주)": 14,

    "평균조회수": 14,
    "평균좋아요": 14,
    "평균댓글": 12,

    "조회수/팔로워(%)": 16,
    "좋아요율(%)": 12,
    "댓글율(%)": 12,
    "공개참여율(ER)": 14,
    "Loyalty Score": 14,

    "댓글작성자중복률(%)": 18,
    "고정댓글러": 12,
    "평균댓글길이": 14,

    "집계기준": 18,
    "계산일시": 18,
}

for header, width in width_map.items():
    col_letter = header_to_col.get(header)
    if col_letter:
        ws.column_dimensions[col_letter].width = width

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(filename)


# =====================================
# 요약
# =====================================
print("=" * 60)
print(f"완료 : {filename}")
print(f"채널 수 : {len(df):,}")

n_metric = (df["표본수"] != "N/A").sum()
n_l3 = (df["평균댓글길이"] != "N/A").sum()
print(f"L2 지표 계산됨 : {n_metric:,}개")
print(f"L3 지표 계산됨 : {n_l3:,}개")
if n_metric < len(df):
    print(f"⚠️ 미계산 {len(df) - n_metric:,}개 "
          f"— 3개월 영상 {getattr(config, 'MIN_SAMPLE', 10)}개 미만이거나 "
          f"activity_status 미해당")
print("ℹ️ Loyalty Score 는 팔로워로 나눈 값(×100 아님) — 타 플랫폼과 비교 불가")
print("=" * 60)