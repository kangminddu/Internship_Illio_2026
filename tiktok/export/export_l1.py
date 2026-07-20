# tiktok/export_l1.py
# DB의 틱톡 L1(채널+최신 snapshot) → 엑셀
#   python -m tiktok.export_l1           # 성공 + 실패(not_found/duplicate)
#   python -m tiktok.export_l1 --all     # + unresolved(단축링크)까지 전체

import argparse
import os
from datetime import datetime

try:
    import pymysql
except ImportError:
    pymysql = None

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from tiktok import config


HEADERS = [
    "번호",
    "creator_id",
    "키값",
    "크리에이터",
    "소속",
    "채널명",
    "핸들URL",
    "수집결과",
    "실패사유",
    "팔로워",
    "팔로잉",
    "총영상수",
    "누적좋아요",
    "Bio",
    "외부링크",
    "sec_uid",
    "수집시각",
]


# 수집결과 / 실패사유는 channel_id_status로 판별.
# (TikTok L1은 crawl_logs를 안 남기므로 상태값으로 대체)
QUERY = """
SELECT
    ROW_NUMBER() OVER (
        ORDER BY (s.follower_count IS NULL), s.follower_count DESC
    ) AS num,
    c.creator_id,
    cr.seed_key,
    cr.nickname,
    COALESCE(cr.agency, '-') AS agency,
    c.channel_name,
    COALESCE(c.channel_url_normalized, c.channel_url_raw) AS url,
    CASE
        WHEN c.channel_id_status='handle_only' AND c.channel_name IS NOT NULL THEN '성공'
        WHEN c.channel_id_status='not_found'  THEN '실패'
        WHEN c.channel_id_status='duplicate'  THEN '실패'
        WHEN c.channel_id_status='unresolved' THEN '미수집'
        ELSE '미수집'
    END AS result,
    CASE
        WHEN c.channel_id_status='handle_only' AND c.channel_name IS NOT NULL THEN '-'
        WHEN c.channel_id_status='not_found'  THEN '삭제/없는 계정 (TikTok 10221)'
        WHEN c.channel_id_status='duplicate'  THEN '중복 계정'
        WHEN c.channel_id_status='unresolved' THEN '단축링크 미해결 (vt.tiktok.com)'
        WHEN c.channel_name IS NULL           THEN '미수집/파싱실패'
        ELSE '-'
    END AS fail_reason,
    s.follower_count,
    s.following_count,
    s.total_video_count,
    s.total_like_count,
    c.bio,
    c.external_link,
    c.external_channel_id,
    s.captured_at
FROM channels c
JOIN creators cr
    ON c.creator_id = cr.creator_id
LEFT JOIN channel_snapshots s
    ON s.channel_id = c.channel_id
   AND s.snapshot_id = (
        SELECT MAX(snapshot_id)
        FROM channel_snapshots
        WHERE channel_id = c.channel_id
   )
WHERE c.platform='tiktok'
{extra}
ORDER BY (s.follower_count IS NULL), s.follower_count DESC
"""


def fetch_rows(conn, include_all):
    # 기본: 성공 + not_found + duplicate (실패사유가 의미 있는 것들)
    # --all: unresolved(단축링크, 크롤링 시도 안 함)까지 포함
    extra = "" if include_all else "AND c.channel_id_status <> 'unresolved'"
    with conn.cursor() as cur:
        cur.execute(QUERY.format(extra=extra))
        return list(cur.fetchall())


def write_xlsx(rows, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "L1_틱톡"

    ws.append(HEADERS)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # 데이터
    for r in rows:
        out = []
        for v in r:
            if isinstance(v, datetime):
                out.append(v.strftime("%Y-%m-%d %H:%M"))
            elif isinstance(v, str) and v and v[0] in ("=", "+", "-", "@"):
                # 엑셀 수식 오인 방지
                out.append("'" + v)
            else:
                out.append(v)
        ws.append(out)

    # 컬럼 인덱스(1-based): 번호1 creator_id2 키값3 크리에이터4 소속5 채널명6
    # 핸들URL7 수집결과8 실패사유9 팔로워10 팔로잉11 총영상수12 누적좋아요13
    # Bio14 외부링크15 sec_uid16 수집시각17
    num_cols = [10, 11, 12, 13]          # 팔로워~누적좋아요
    center_cols = [1, 2, 8, 9, 17]       # 번호/creator_id/수집결과/실패사유/수집시각

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
        for ci in num_cols:
            cell = row[ci - 1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
            cell.alignment = center
        for ci in center_cols:
            row[ci - 1].alignment = center

    # 열 너비
    widths = [6, 10, 12, 18, 14, 20, 40, 10, 30, 12, 12, 12, 14, 45, 35, 40, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def main():
    ap = argparse.ArgumentParser(prog="tiktok.export_l1")
    ap.add_argument("--all", action="store_true",
                    help="unresolved(단축링크)까지 전체 포함")
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    if pymysql is None:
        raise SystemExit("pip install pymysql")

    out = args.out or os.path.join(
        os.path.dirname(__file__), "output", "L1_틱톡_채널정보.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)

    conn = pymysql.connect(**config.DB)
    try:
        rows = fetch_rows(conn, args.all)
    finally:
        conn.close()

    write_xlsx(rows, out)
    print(f"[export] {len(rows)}행 저장 -> {out}")

    # 검증 — 수집결과 분포
    from collections import Counter
    # result는 8번째 컬럼(index 7)
    dist = Counter(r[7] for r in rows)
    print("수집결과 분포:")
    for k, v in dist.most_common():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()