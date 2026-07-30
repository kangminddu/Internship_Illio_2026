"""
youtube/crawler/seed.py — 파이프라인 진입점

엑셀 한 장을 읽어 creators / channels 두 테이블에 적재한다.
이 파일이 끝나면 channels에는 'URL만 있는 껍데기' 행이 수천 개 생기고,
L1이 그 URL을 하나씩 방문해 나머지를 채운다.

설계 판단 3가지
------
1) 컬럼을 '위치'가 아니라 '헤더 이름'으로 찾는다
   시드 엑셀은 사람이 관리하는 파일이라 컬럼 순서가 수시로 바뀐다.
   row[3]처럼 인덱스로 읽으면 열 하나만 삽입돼도 전부 어긋난다.
   헤더명 + 별칭 사전으로 찾으면 순서가 바뀌어도, 표기가 조금 달라도 견딘다.

2) creators를 먼저 INSERT한다
   channels.creator_id가 FK다. creator 행이 없으면 channel을 넣을 수 없다.
   그리고 한 크리에이터가 유튜브·인스타·틱톡 채널을 각각 가질 수 있으므로
   creator : channel = 1 : N 구조다.

3) URL 정규화를 여기서 미리 한다
   같은 채널이 /@handle, /c/name, /channel/UC... 세 형태로 들어올 수 있다.
   적재 시점에 표준형으로 바꿔두지 않으면 중복 판별이 불가능해진다.
   정규화 로직은 lib/youtube_url_filter.py 한 곳에 몰아뒀다.

사용법:
  python -m youtube.crawler.seed --file 파일.xlsx
  python -m youtube.crawler.seed --file 파일.xlsx --sheet "시트명"
"""
import re
import argparse
from datetime import datetime
from openpyxl import load_workbook
import pymysql
from youtube.config import DB
from youtube.crawler.lib.youtube_url_filter import normalize_youtube_channel_url


# ─────────────────────────────────────────────
# 컬럼 별칭 사전
#
# 엑셀을 만드는 사람마다 헤더를 다르게 쓴다.
# "유튜브" / "YouTube URL" / "유튜브 url" 이 전부 같은 열이다.
# 여기 후보를 나열해두면 어느 표기로 와도 잡힌다.
# 새 엑셀에서 컬럼을 못 찾으면 실제 헤더명을 이 리스트에 추가하면 된다.
# ─────────────────────────────────────────────
COLUMN_ALIASES = {
    "youtube":  ["유튜브", "youtubeurl", "youtube", "유튜브url", "youtube_url", "유튜브 url"],
    "nickname": ["닉네임", "상품명2", "상품명 2", "이름", "크리에이터", "키값"],
    "agency":   ["소속사", "소속", "agency"],
    "category": ["구분", "카테고리", "category"],
    "birthday": ["생일", "birthday"],
    "debut":    ["데뷔", "데뷔일", "debut"],
}


def normalize_header(h):
    """헤더 비교용 정규화: 공백 제거 + 소문자.
    "YouTube URL"과 "youtubeurl"을 같은 것으로 취급하기 위함."""
    if h is None:
        return ""
    return str(h).strip().replace(" ", "").lower()


def build_column_map(header_row):
    """헤더 행 → {필드명: 열 인덱스} 매핑.

    별칭 리스트를 앞에서부터 훑어 처음 매칭되는 것을 쓴다.
    못 찾은 필드는 colmap에 아예 없고, get()이 None을 반환한다.
    """
    norm_headers = [normalize_header(h) for h in header_row]
    colmap = {}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            na = normalize_header(alias)
            if na in norm_headers:
                colmap[field] = norm_headers.index(na)
                break
    return colmap


def classify_status(u):
    """URL 형태로 channel_id_status를 정한다.

    이 값이 곧 "이 URL로 채널을 특정할 수 있는가"의 답이다.

      resolved     /channel/UC...  → 유튜브 고유 ID 확보. 확정.
      handle_only  /@name          → 핸들. 바뀔 수 있지만 접근 가능.
      custom_only  /c/name         → 구형 커스텀 URL.
      user_legacy  /user/name      → 더 오래된 형식.
      unresolved   그 외           → 형태를 판별 못 함.

    resolved 외에는 전부 '아직 UC를 모르는 상태'다.
    L1이 실제로 페이지를 열어 UC를 찾아내면 resolved로 승격시킨다.
    (crawler_l1_parallel.save_result에서 UPDATE)

    구분해두는 이유: 나중에 "핸들이 바뀌어 못 찾은 채널"과
    "애초에 URL이 이상했던 채널"을 분리해서 볼 수 있다.
    """
    if re.search(r"/channel/UC[\w-]{22}", u): return "resolved"
    if "/@" in u: return "handle_only"
    if "/c/" in u: return "custom_only"
    if "/user/" in u: return "user_legacy"
    return "unresolved"   # 구형 custom URL(youtube.com/이름)도 여기로 → 크롤링 시 UC로 해소


def clean(v):
    """엑셀 셀 → 문자열 or None.
    사람이 채운 엑셀에는 빈칸 대신 "-"나 "NULL"이 들어있는 경우가 많다."""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "NULL"):
        return None
    return s


def to_date(v):
    """openpyxl은 날짜 서식 셀을 datetime으로 준다.
    문자열로 적힌 날짜("2020.03.01" 등)는 형식이 제각각이라 버린다."""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return None


def get(row, colmap, field):
    """colmap을 통해 안전하게 셀 값 읽기.

    idx >= len(row) 체크가 필요한 이유:
    엑셀 행마다 길이가 다를 수 있다. 뒤쪽 열이 전부 비어있으면
    openpyxl이 그 행을 짧게 반환해서 IndexError가 난다.
    """
    idx = colmap.get(field)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def main(xlsx_path, sheet_name=None, key_prefix="SD"):
    wb = load_workbook(xlsx_path, read_only=True)   # read_only: 대용량 엑셀 메모리 절약

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
        if sheet_name:
            print(f"⚠️  시트 '{sheet_name}' 없음 → 첫 시트 '{ws.title}' 사용")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("❌ 빈 시트")
        return

    header = rows[0]
    colmap = build_column_map(header)

    # 컬럼 인식 결과를 먼저 출력한다.
    # 수천 행을 넣고 나서 "엉뚱한 열을 읽었다"를 발견하면 늦다.
    # 필수 컬럼이 없으면 DB를 건드리기 전에 중단한다.
    print(f"시트: {ws.title} | 데이터 {len(rows)-1}행")
    print(f"인식된 컬럼: {colmap}")
    if "youtube" not in colmap:
        print("❌ '유튜브' 컬럼을 못 찾음. 헤더:", header)
        print("   → COLUMN_ALIASES['youtube']에 실제 헤더명을 추가하세요.")
        return
    if "nickname" not in colmap:
        print("❌ '닉네임' 컬럼을 못 찾음. 헤더:", header)
        return
    print()

    conn = pymysql.connect(**DB, autocommit=True)
    inserted_ch = 0
    creator_only = 0
    skipped = 0
    skip_reasons = {}   # 사유별 집계 (마지막에 요약 출력)

    with conn.cursor() as cur:
        for i, row in enumerate(rows[1:], 1):
            nickname = clean(get(row, colmap, "nickname"))
            youtube  = clean(get(row, colmap, "youtube"))
            agency   = clean(get(row, colmap, "agency"))
            category = clean(get(row, colmap, "category"))
            debut    = to_date(get(row, colmap, "debut"))
            birthday = to_date(get(row, colmap, "birthday"))

            if not nickname:
                skipped += 1
                continue

            # seed_key = 엑셀 행 번호 기반 고유키.
            # 이게 있어야 같은 엑셀을 다시 돌려도 중복 생성되지 않는다(멱등성).
            # creators.seed_key에 UNIQUE가 걸려 있어 ON DUPLICATE가 동작한다.
            seed_key = f"{key_prefix}_{i}"

            # ── creator 먼저 (channels.creator_id가 FK이므로) ──
            # ON DUPLICATE KEY UPDATE: 재실행 시 엑셀의 최신 값으로 갱신한다.
            # 소속사가 바뀌거나 닉네임이 수정된 경우를 반영하기 위함.
            cur.execute("""
                INSERT INTO creators (seed_key, nickname, category, agency, debut_date, birthday)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    nickname=VALUES(nickname), category=VALUES(category),
                    agency=VALUES(agency), debut_date=VALUES(debut_date),
                    birthday=VALUES(birthday)
            """, (seed_key, nickname, category, agency, debut, birthday))

            # 유튜브 URL이 없는 크리에이터도 creators에는 남긴다.
            # 나중에 인스타/틱톡 시드가 같은 seed_key로 채널을 붙일 수 있다.
            if not youtube:
                creator_only += 1
                continue

            # ── URL 검증 + 정규화 (한 번에) ──
            # percent-encoded 한글 핸들, /videos 같은 탭 꼬리, 오타 도메인 등을
            # 여기서 걸러낸다. 실패하면 사유를 받아 집계한다.
            url_norm, skip_reason = normalize_youtube_channel_url(youtube)
            if not url_norm:
                print(f"⚠️ Skip [{skip_reason}] : {youtube}")
                skip_reasons[skip_reason] = skip_reasons.get(skip_reason, 0) + 1
                creator_only += 1
                continue

            # 방금 INSERT/UPDATE한 creator의 PK를 가져온다.
            # (ON DUPLICATE 경로에서는 lastrowid가 신뢰할 수 없어 직접 조회)
            cur.execute("SELECT creator_id FROM creators WHERE seed_key=%s", (seed_key,))
            creator_id = cur.fetchone()[0]

            status = classify_status(url_norm)

            # URL에 UC ID가 이미 들어있으면 미리 뽑아둔다.
            # 있으면 L1이 굳이 찾을 필요가 없고, 중복 판별도 바로 된다.
            uc = None
            m = re.search(r"(UC[\w-]{22})", url_norm)
            if m:
                uc = m.group(1)

            # channel_url_raw에 UNIQUE가 걸려 있어 같은 URL은 한 번만 들어간다.
            # 재실행 시엔 정규화 결과와 상태만 갱신한다.
            cur.execute("""
                INSERT INTO channels
                  (creator_id, platform, channel_url_raw, channel_url_normalized,
                   channel_id_status, external_channel_id, is_primary)
                VALUES (%s, 'youtube', %s, %s, %s, %s, 1)
                ON DUPLICATE KEY UPDATE
                  channel_url_normalized=VALUES(channel_url_normalized),
                  channel_id_status=VALUES(channel_id_status)
            """, (creator_id, youtube, url_norm, status, uc))
            inserted_ch += 1

    conn.close()

    # 요약 출력.
    # 특히 skip 사유별 집계가 중요하다 — 시드 엑셀의 품질을 보여준다.
    # (예: "단축링크 15건", "도메인 오타 3건" → 엑셀 관리자에게 피드백 가능)
    print(f"\ncreators 총: {inserted_ch + creator_only}")
    print(f"  ├ 유튜브 채널 생성: {inserted_ch}")
    print(f"  └ creators만(유튜브 없음/부적합): {creator_only}")
    print(f"닉네임 없어 skip: {skipped}")
    if skip_reasons:
        print("URL skip 사유별 집계:")
        for reason, cnt in sorted(skip_reasons.items(), key=lambda x: -x[1]):
            print(f"  - {reason}: {cnt}건")
    print("완료. python main.py --l1 로 시작.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="엑셀 → DB seed")
    parser.add_argument("--file", required=True, help="엑셀 파일 경로")
    parser.add_argument("--sheet", default=None, help="시트명 (없으면 첫 시트)")
    parser.add_argument("--prefix", default="SD", help="seed_key 접두사 (기본 SD)")
    args = parser.parse_args()
    main(args.file, args.sheet, args.prefix)