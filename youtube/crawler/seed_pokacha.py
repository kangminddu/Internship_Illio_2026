"""
포카챠 섭외 통합 DB → creators/channels seed.
유튜브 URL 있는 셀럽만 (988명). 닉네임/소속/데뷔/생일/구분까지 적재.
ON DUPLICATE KEY UPDATE로 재실행 안전.
"""
import re
from datetime import datetime
from openpyxl import load_workbook
import pymysql
from youtube.config import DB

XLSX = "/Users/kangminsoo/Desktop/Internship_Illio_2026/포카챠_섭외_DB_필터_20260708.xlsx"
SHEET = "섭외 통합 DB"


def normalize_url(url):
    if not url:
        return None
    url = str(url).strip()
    if not url.startswith("http"):
        url = "https://" + url
    url = re.sub(r"\?.*$", "", url)
    for suffix in ("/featured", "/videos", "/about", "/discussion",
                   "/community", "/playlists", "/streams", "/shorts"):
        if url.endswith(suffix):
            url = url[:-len(suffix)]
    return url.rstrip("/")


def classify_status(u):
    if re.search(r"/channel/UC[\w-]{22}", u): return "resolved"
    if "/@" in u: return "handle_only"
    if "/c/" in u: return "custom_only"
    if "/user/" in u: return "user_legacy"
    return "unresolved"


def clean(v):
    """엑셀 셀 정리. '-', 'NULL', 빈칸 → None."""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "NULL"):
        return None
    return s


def to_date(v):
    """datetime 셀 → 'YYYY-MM-DD'. 생일의 1900년은 연도 무의미하나 그대로 저장."""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return None


def main():
    wb = load_workbook(XLSX, read_only=True)
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))[1:]  # 헤더 제외

    conn = pymysql.connect(**DB, autocommit=True)
    inserted_ch = 0
    creator_only = 0
    skipped = 0

    with conn.cursor() as cur:
        for i, r in enumerate(rows, 1):
            gubun    = clean(r[0])   # 구분 (버츄얼 등) → category
            nickname = clean(r[1])   # 닉네임
            youtube  = clean(r[5])   # 유튜브
            agency   = clean(r[8])   # 소속사
            debut    = to_date(r[10])  # 데뷔
            birthday = to_date(r[11])  # 생일

            if not nickname:
                skipped += 1
                continue

            seed_key = f"PK_{i}"

            # creators (닉네임/소속/카테고리/데뷔/생일)
            cur.execute("""
                INSERT INTO creators (seed_key, nickname, category, agency, debut_date, birthday)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    nickname=VALUES(nickname), category=VALUES(category),
                    agency=VALUES(agency), debut_date=VALUES(debut_date),
                    birthday=VALUES(birthday)
            """, (seed_key, nickname, gubun, agency, debut, birthday))

            # 유튜브 URL 없으면 creators만 (channels 안 만듦)
            if not youtube:
                creator_only += 1
                continue

            cur.execute("SELECT creator_id FROM creators WHERE seed_key=%s", (seed_key,))
            creator_id = cur.fetchone()[0]

            url_norm = normalize_url(youtube)
            status = classify_status(youtube)
            uc = None
            m = re.search(r"(UC[\w-]{22})", youtube)
            if m:
                uc = m.group(1)

            cur.execute("""
                INSERT INTO channels
                  (creator_id, platform, channel_url_raw, channel_url_normalized,
                   channel_id_status, external_channel_id, is_primary)
                VALUES (%s, 'youtube', %s, %s, %s, %s, 1)
                ON DUPLICATE KEY UPDATE
                  channel_url_normalized=VALUES(channel_url_normalized),
                  channel_id_status=VALUES(channel_id_status)
            """, (creator_id, youtube, url_norm, status, uc))
            inserted_ch += 1

    conn.close()
    print(f"creators 총: {inserted_ch + creator_only}")
    print(f"  ├ 유튜브 채널 생성: {inserted_ch}")
    print(f"  └ creators만(유튜브 없음): {creator_only}")
    print(f"닉네임 없어 skip: {skipped}")
    print("완료. python main.py --l1 로 시작.")


if __name__ == "__main__":
    main()