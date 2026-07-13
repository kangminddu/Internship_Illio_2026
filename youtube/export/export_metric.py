import os
import warnings
warnings.filterwarnings("ignore", category=UserWarning)
import pandas as pd
import pymysql

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import DB, EXPORT_DIR

# =====================================
# SQL — 모든 파생지표를 channel_metrics(m)에서 직접 (export 재계산 없음)
# =====================================
sql = """
SELECT
    ROW_NUMBER() OVER (ORDER BY cr.nickname) AS 번호,
    cr.nickname AS 크리에이터,
    COALESCE(cr.agency,'-') AS 소속,

    MAX(chs.follower_count) AS follower_count,

    -- 통합 지표
    MAX(m.view_per_follower_ratio) AS vpf,
    MAX(m.engagement_rate)         AS er,
    MAX(m.loyalty_score)           AS loyalty,
    MAX(m.upload_frequency_weekly) AS upload_freq,
    MAX(m.avg_view)                AS total_avg_view,

    -- 최근 3개월
    MAX(m.videos_3m)          AS videos_3m,
    MAX(m.avg_view_3m)        AS avg_view_3m,
    MAX(m.avg_like_3m)        AS avg_like_3m,
    MAX(m.avg_comment_3m)     AS avg_comment_3m,
    MAX(m.engagement_rate_3m) AS engagement_rate_3m,

    -- 롱폼
    MAX(m.longform_avg_view)    AS longform_avg_view,
    MAX(m.longform_avg_like)    AS longform_avg_like,
    MAX(m.longform_avg_comment) AS longform_avg_comment,
    MAX(m.longform_er)          AS longform_er,

    -- 쇼츠
    MAX(m.shorts_avg_view)    AS shorts_avg_view,
    MAX(m.shorts_avg_like)    AS shorts_avg_like,
    MAX(m.shorts_avg_comment) AS shorts_avg_comment,
    MAX(m.shorts_er)          AS shorts_er,

    -- 광고 롱폼
    MAX(m.ad_longform_avg_view)    AS ad_longform_avg_view,
    MAX(m.ad_longform_avg_like)    AS ad_longform_avg_like,
    MAX(m.ad_longform_avg_comment) AS ad_longform_avg_comment,
    MAX(m.ad_longform_er)          AS ad_longform_er,

    -- 일반 롱폼
    MAX(m.normal_longform_avg_view)    AS normal_longform_avg_view,
    MAX(m.normal_longform_avg_like)    AS normal_longform_avg_like,
    MAX(m.normal_longform_avg_comment) AS normal_longform_avg_comment,
    MAX(m.normal_longform_er)          AS normal_longform_er,

    -- 광고 쇼츠
    MAX(m.ad_shorts_avg_view)    AS ad_shorts_avg_view,
    MAX(m.ad_shorts_avg_like)    AS ad_shorts_avg_like,
    MAX(m.ad_shorts_avg_comment) AS ad_shorts_avg_comment,
    MAX(m.ad_shorts_er)          AS ad_shorts_er,

    -- 일반 쇼츠
    MAX(m.normal_shorts_avg_view)    AS normal_shorts_avg_view,
    MAX(m.normal_shorts_avg_like)    AS normal_shorts_avg_like,
    MAX(m.normal_shorts_avg_comment) AS normal_shorts_avg_comment,
    MAX(m.normal_shorts_er)          AS normal_shorts_er,

    -- L3 (댓글)
    MAX(m.commenter_overlap_rate)  AS commenter_overlap_rate,
    MAX(m.regular_commenter_count) AS regular_commenter_count,
    MAX(m.avg_comment_length)      AS avg_comment_length

FROM creators cr
JOIN channels ch
    ON cr.creator_id = ch.creator_id
LEFT JOIN channel_snapshots chs
    ON ch.channel_id = chs.channel_id
   AND chs.captured_at = (
        SELECT MAX(captured_at) FROM channel_snapshots
        WHERE channel_id = ch.channel_id
   )
LEFT JOIN channel_metrics m
    ON ch.channel_id = m.channel_id
GROUP BY ch.channel_id
ORDER BY cr.nickname;
"""

# =====================================
# SQL → DataFrame
# =====================================
conn = pymysql.connect(**DB)
df = pd.read_sql(sql, conn)
conn.close()
print(df.head())
print(df.columns)

# =====================================
# 컬럼 매핑 (계산 없이 calc_metrics 값 그대로)
# =====================================
df["구독자대비조회율(%)"] = df["vpf"]
df["공개참여율(ER)"]      = df["er"]
df["업로드빈도(주)"]      = df["upload_freq"]
df["Loyalty Score"]      = df["loyalty"]

df["최근3개월업로드"]     = df["videos_3m"]
df["최근3개월평균조회수"]  = df["avg_view_3m"]
df["최근3개월평균좋아요"]  = df["avg_like_3m"]
df["최근3개월평균댓글"]    = df["avg_comment_3m"]
df["최근3개월ER"]         = df["engagement_rate_3m"]

df["전체평균조회수"]      = df["total_avg_view"]

# 롱폼
df["롱폼평균조회수"] = df["longform_avg_view"]
df["롱폼평균좋아요"] = df["longform_avg_like"]
df["롱폼평균댓글"] = df["longform_avg_comment"]
df["롱폼ER"] = df["longform_er"]

# 쇼츠
df["쇼츠평균조회수"] = df["shorts_avg_view"]
df["쇼츠평균좋아요"] = df["shorts_avg_like"]
df["쇼츠평균댓글"] = df["shorts_avg_comment"]
df["쇼츠ER"] = df["shorts_er"]

# 광고 롱폼
df["광고롱폼평균조회수"] = df["ad_longform_avg_view"]
df["광고롱폼평균좋아요"] = df["ad_longform_avg_like"]
df["광고롱폼평균댓글"] = df["ad_longform_avg_comment"]
df["광고롱폼ER"] = df["ad_longform_er"]

# 일반 롱폼
df["일반롱폼평균조회수"] = df["normal_longform_avg_view"]
df["일반롱폼평균좋아요"] = df["normal_longform_avg_like"]
df["일반롱폼평균댓글"] = df["normal_longform_avg_comment"]
df["일반롱폼ER"] = df["normal_longform_er"]

# 광고 쇼츠
df["광고쇼츠평균조회수"] = df["ad_shorts_avg_view"]
df["광고쇼츠평균좋아요"] = df["ad_shorts_avg_like"]
df["광고쇼츠평균댓글"] = df["ad_shorts_avg_comment"]
df["광고쇼츠ER"] = df["ad_shorts_er"]

# 일반 쇼츠
df["일반쇼츠평균조회수"] = df["normal_shorts_avg_view"]
df["일반쇼츠평균좋아요"] = df["normal_shorts_avg_like"]
df["일반쇼츠평균댓글"] = df["normal_shorts_avg_comment"]
df["일반쇼츠ER"] = df["normal_shorts_er"]

df["댓글작성자중복률(%)"]  = df["commenter_overlap_rate"]
df["고정댓글러"]          = df["regular_commenter_count"]
df["평균댓글길이"]        = df["avg_comment_length"]

# =====================================
# NaN → N/A
# =====================================
na_cols = [
    "번호","크리에이터","소속",

    "구독자대비조회율(%)",
    "공개참여율(ER)",
    "업로드빈도(주)",
    "Loyalty Score",

    "최근3개월업로드",
    "최근3개월평균조회수",
    "최근3개월평균좋아요",
    "최근3개월평균댓글",
    "최근3개월ER",

    "전체평균조회수",

    "롱폼평균조회수",
    "롱폼평균좋아요",
    "롱폼평균댓글",
    "롱폼ER",

    "쇼츠평균조회수",
    "쇼츠평균좋아요",
    "쇼츠평균댓글",
    "쇼츠ER",

    "광고롱폼평균조회수",
    "광고롱폼평균좋아요",
    "광고롱폼평균댓글",
    "광고롱폼ER",

    "일반롱폼평균조회수",
    "일반롱폼평균좋아요",
    "일반롱폼평균댓글",
    "일반롱폼ER",

    "광고쇼츠평균조회수",
    "광고쇼츠평균좋아요",
    "광고쇼츠평균댓글",
    "광고쇼츠ER",

    "일반쇼츠평균조회수",
    "일반쇼츠평균좋아요",
    "일반쇼츠평균댓글",
    "일반쇼츠ER",

    "댓글작성자중복률(%)",
    "고정댓글러",
    "평균댓글길이",
]
for col in na_cols:
    df[col] = df[col].apply(lambda x: "N/A" if pd.isna(x) else x)

# =====================================
# 최종 컬럼 순서 (표 흐름대로)
# =====================================
df = df[[
    "번호",
    "크리에이터",
    "소속",

    "구독자대비조회율(%)",
    "공개참여율(ER)",
    "업로드빈도(주)",
    "Loyalty Score",

    "최근3개월업로드",
    "최근3개월평균조회수",
    "최근3개월평균좋아요",
    "최근3개월평균댓글",
    "최근3개월ER",

    "전체평균조회수",

    "롱폼평균조회수",
    "롱폼평균좋아요",
    "롱폼평균댓글",
    "롱폼ER",

    "쇼츠평균조회수",
    "쇼츠평균좋아요",
    "쇼츠평균댓글",
    "쇼츠ER",

    "광고롱폼평균조회수",
    "광고롱폼평균좋아요",
    "광고롱폼평균댓글",
    "광고롱폼ER",

    "일반롱폼평균조회수",
    "일반롱폼평균좋아요",
    "일반롱폼평균댓글",
    "일반롱폼ER",

    "광고쇼츠평균조회수",
    "광고쇼츠평균좋아요",
    "광고쇼츠평균댓글",
    "광고쇼츠ER",

    "일반쇼츠평균조회수",
    "일반쇼츠평균좋아요",
    "일반쇼츠평균댓글",
    "일반쇼츠ER",

    "댓글작성자중복률(%)",
    "고정댓글러",
    "평균댓글길이",
]]

# =====================================
# Excel 저장
# =====================================
filename = os.path.join(EXPORT_DIR, "파생지표_포카챠.xlsx")
df.to_excel(filename, index=False, sheet_name="Metrics")

wb = load_workbook(filename)
ws = wb["Metrics"]

# =====================================
# 스타일
# =====================================
header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF")
center = Alignment(horizontal="center", vertical="center")
thin = Side(border_style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = center

# =====================================
# 숫자 서식 (헤더명 기반 — 순서 바뀌어도 안 깨짐)
# =====================================
fmt_map = {
    "구독자대비조회율(%)":"0.00",
    "공개참여율(ER)":"0.00",
    "업로드빈도(주)":"0.00",
    "Loyalty Score":"0.0000",

    "최근3개월업로드":"#,##0",
    "최근3개월평균조회수":"#,##0.00",
    "최근3개월평균좋아요":"#,##0.00",
    "최근3개월평균댓글":"#,##0.00",
    "최근3개월ER":"0.00",

    "전체평균조회수":"#,##0.00",

    "롱폼평균조회수":"#,##0.00",
    "롱폼평균좋아요":"#,##0.00",
    "롱폼평균댓글":"#,##0.00",
    "롱폼ER":"0.00",

    "쇼츠평균조회수":"#,##0.00",
    "쇼츠평균좋아요":"#,##0.00",
    "쇼츠평균댓글":"#,##0.00",
    "쇼츠ER":"0.00",

    "광고롱폼평균조회수":"#,##0.00",
    "광고롱폼평균좋아요":"#,##0.00",
    "광고롱폼평균댓글":"#,##0.00",
    "광고롱폼ER":"0.00",

    "일반롱폼평균조회수":"#,##0.00",
    "일반롱폼평균좋아요":"#,##0.00",
    "일반롱폼평균댓글":"#,##0.00",
    "일반롱폼ER":"0.00",

    "광고쇼츠평균조회수":"#,##0.00",
    "광고쇼츠평균좋아요":"#,##0.00",
    "광고쇼츠평균댓글":"#,##0.00",
    "광고쇼츠ER":"0.00",

    "일반쇼츠평균조회수":"#,##0.00",
    "일반쇼츠평균좋아요":"#,##0.00",
    "일반쇼츠평균댓글":"#,##0.00",
    "일반쇼츠ER":"0.00",

    "댓글작성자중복률(%)":"0.00",
    "고정댓글러":"#,##0",
    "평균댓글길이":"0.00",
}
header_to_col = {}
for idx, cell in enumerate(ws[1], start=1):
    header_to_col[cell.value] = get_column_letter(idx)

for header, fmt in fmt_map.items():
    col_letter = header_to_col.get(header)
    if not col_letter:
        continue
    for cell in ws[col_letter][1:]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = fmt

# =====================================
# 열 너비 (헤더명 기반)
# =====================================
width_map = {
    "번호":8,
    "크리에이터":22,
    "소속":18,

    "구독자대비조회율(%)":16,
    "공개참여율(ER)":14,
    "업로드빈도(주)":14,
    "Loyalty Score":14,

    "최근3개월업로드":14,
    "최근3개월평균조회수":16,
    "최근3개월평균좋아요":16,
    "최근3개월평균댓글":16,
    "최근3개월ER":12,

    "전체평균조회수":16,

    "롱폼평균조회수":16,
    "롱폼평균좋아요":16,
    "롱폼평균댓글":16,
    "롱폼ER":12,

    "쇼츠평균조회수":16,
    "쇼츠평균좋아요":16,
    "쇼츠평균댓글":16,
    "쇼츠ER":12,

    "광고롱폼평균조회수":18,
    "광고롱폼평균좋아요":18,
    "광고롱폼평균댓글":18,
    "광고롱폼ER":14,

    "일반롱폼평균조회수":18,
    "일반롱폼평균좋아요":18,
    "일반롱폼평균댓글":18,
    "일반롱폼ER":14,

    "광고쇼츠평균조회수":18,
    "광고쇼츠평균좋아요":18,
    "광고쇼츠평균댓글":18,
    "광고쇼츠ER":14,

    "일반쇼츠평균조회수":18,
    "일반쇼츠평균좋아요":18,
    "일반쇼츠평균댓글":18,
    "일반쇼츠ER":14,

    "댓글작성자중복률(%)":18,
    "고정댓글러":12,
    "평균댓글길이":14,
}
for header, width in width_map.items():
    col_letter = header_to_col.get(header)
    if col_letter:
        ws.column_dimensions[col_letter].width = width

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(filename)

print("=" * 60)
print(f"완료 : {filename}")
print(f"채널 수 : {len(df)}")
print("=" * 60)
