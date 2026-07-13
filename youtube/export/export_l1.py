import os
import warnings
warnings.filterwarnings("ignore", category=UserWarning)
import pandas as pd
import pymysql

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import DB, EXPORT_DIR

sql = """
SELECT
    ROW_NUMBER() OVER (
        ORDER BY cs.follower_count DESC
    ) AS '번호',
    cr.nickname AS '크리에이터',
    COALESCE(cr.agency,'-') AS '소속',
    CASE
        WHEN ch.channel_activity_status='active'     THEN '활동중'
        WHEN ch.channel_activity_status='low_active' THEN '저활동'
        WHEN ch.channel_activity_status='inactive'   THEN '휴면'
        ELSE '-'
    END AS '활동상태',
    cs.follower_count AS '구독자수',
    cs.total_view_count AS '누적조회수',
    cs.total_video_count AS '총영상수',
    DATE_FORMAT(ch.channel_opened_at,'%Y-%m-%d') AS '채널개설일',
    CASE
        WHEN log.status='success' THEN '성공'
        WHEN log.status='failed'  THEN '실패'
        ELSE '-'
    END AS '수집결과',
    COALESCE(log.error_type,'-')  AS '실패사유',
    COALESCE(log.http_status,'-') AS 'HTTP코드',
    COALESCE(ch.channel_url_normalized, ch.channel_url_raw) AS 'URL'
FROM creators cr
JOIN channels ch
    ON cr.creator_id = ch.creator_id
LEFT JOIN (
    SELECT cs1.*
    FROM channel_snapshots cs1
    JOIN (
        SELECT channel_id, MAX(captured_at) latest
        FROM channel_snapshots
        GROUP BY channel_id
    ) t
    ON cs1.channel_id=t.channel_id AND cs1.captured_at=t.latest
) cs
    ON ch.channel_id=cs.channel_id
LEFT JOIN (
    SELECT c.channel_id, c.status, c.error_type, c.http_status
    FROM crawl_logs c
    JOIN (
        SELECT channel_id, MAX(log_id) AS mx
        FROM crawl_logs
        WHERE layer='L1' AND channel_id IS NOT NULL
        GROUP BY channel_id
    ) m ON c.log_id = m.mx
) log
    ON log.channel_id = ch.channel_id
ORDER BY cs.follower_count DESC;
"""


def sanitize_formula(val):
    """엑셀이 수식으로 오인하는 것 방지: =+-@ 로 시작하면 앞에 ' 붙임"""
    if isinstance(val, str) and val and val[0] in ('=', '+', '-', '@'):
        return "'" + val
    return val


# ----------------------------------------
# DB → DataFrame
# ----------------------------------------
conn = pymysql.connect(**DB)
df = pd.read_sql(sql, conn)
conn.close()

# 텍스트 컬럼 수식 오인 방지 (df 생성 후 적용)
for col in ["크리에이터", "소속", "URL"]:
    df[col] = df[col].apply(sanitize_formula)

filename = os.path.join(EXPORT_DIR, "L1_채널정보_포카챠.xlsx")
df.to_excel(filename, index=False, sheet_name="L1")

# ----------------------------------------
# Excel 꾸미기
# ----------------------------------------
wb = load_workbook(filename)
ws = wb["L1"]

header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF")
center = Alignment(horizontal="center", vertical="center")
thin = Side(border_style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        if cell.column in [1, 5, 6, 7]:
            cell.alignment = center

for col in ["E", "F", "G"]:
    for cell in ws[col][1:]:
        cell.number_format = '#,##0'

for column_cells in ws.columns:
    length = max(
        len(str(cell.value)) if cell.value is not None else 0
        for cell in column_cells
    )
    letter = get_column_letter(column_cells[0].column)
    ws.column_dimensions[letter].width = min(length + 4, 35)

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(filename)

print(f"완료 : {filename}")
print(f"총 {len(df)}개 채널")

# 검증 — 활동상태 분포 (DB 최신값 반영됐는지)
print("활동상태 분포:")
print(df['활동상태'].value_counts().to_string())