import os
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

import pandas as pd
import pymysql

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from instagram.config import DB, OUTPUT_DIR


# ⚠️ DATE_FORMAT 의 리터럴 % 는 pymysql 이 플레이스홀더로 오인한다.
#    포맷은 엑셀 number_format 에 맡기고 여기서는 datetime 그대로 넘긴다.
sql = """
SELECT

ROW_NUMBER() OVER(
ORDER BY
cr.nickname,
ct.published_at DESC,
c.published_at DESC
) AS '번호',

cr.nickname AS '크리에이터',

CASE
    WHEN ct.content_type='feed_image' THEN '피드'
    WHEN ct.content_type='carousel' THEN '캐러셀'
    WHEN ct.content_type='reels' THEN '릴스'
    ELSE ct.content_type
END AS '콘텐츠유형',

DATE(ct.published_at) AS '게시일',

ct.external_id AS '포스트ID',

c.external_comment_id AS '댓글ID',

f.external_author_id AS '댓글작성자ID',

c.author_display_name AS '댓글작성자',

c.published_at AS '댓글작성일',

c.like_count AS '댓글좋아요',

CHAR_LENGTH(c.comment_text) AS '댓글길이',

c.comment_text AS '댓글내용',

-- 릴스는 /p/ 로 열면 /reels/ 로 리다이렉트되므로 유형별로 분기
CONCAT(
'https://www.instagram.com/',
CASE WHEN ct.content_type='reels' THEN 'reel/' ELSE 'p/' END,
ct.external_id,
'/'
) AS '포스트URL'

FROM comments c

JOIN fans f
ON c.fan_id = f.fan_id

JOIN contents ct
ON c.content_id = ct.content_id

JOIN channels ch
ON ct.channel_id = ch.channel_id

JOIN creators cr
ON ch.creator_id = cr.creator_id

WHERE ch.platform = 'instagram'

ORDER BY
cr.nickname,
ct.published_at DESC,
c.published_at DESC;
"""


def sanitize_formula(val):
    """엑셀이 수식으로 오인하는 것 방지"""
    if isinstance(val, str) and val and val[0] in ("=", "+", "-", "@"):
        return "'" + val
    return val


# ------------------------------------------------
# DB -> DataFrame
# ------------------------------------------------
conn = pymysql.connect(**DB)
try:
    df = pd.read_sql(sql, conn)
finally:
    conn.close()


# ------------------------------------------------
# 수식 오인 방지
# ------------------------------------------------
for col in ["댓글내용", "댓글작성자", "크리에이터"]:
    if col in df.columns:
        df[col] = df[col].fillna("").apply(sanitize_formula)


# ------------------------------------------------
# Excel 저장
# ------------------------------------------------
os.makedirs(OUTPUT_DIR, exist_ok=True)
filename = os.path.join(
    OUTPUT_DIR,
    "L3_인스타그램_댓글정보.xlsx"
)

df.to_excel(
    filename,
    index=False,
    sheet_name="L3"
)


# ------------------------------------------------
# Excel Styling
# ------------------------------------------------
wb = load_workbook(filename)
ws = wb["L3"]

header_fill = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
)

header_font = Font(
    bold=True,
    color="FFFFFF"
)

center = Alignment(
    horizontal="center",
    vertical="center"
)

left = Alignment(
    horizontal="left",
    vertical="center"
)

thin = Side(
    border_style="thin",
    color="D9D9D9"
)

border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)

# -------------------------
# Header
# -------------------------
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

# -------------------------
# Body
# -------------------------
# 가운데 정렬: 번호/유형/게시일/댓글작성일/좋아요/길이
CENTER_COLS = {1, 3, 4, 9, 10, 11}

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = center if cell.column in CENTER_COLS else left


# -------------------------
# Number / Date Format
# -------------------------
for col in ["J", "K"]:               # 댓글좋아요 / 댓글길이
    for cell in ws[col][1:]:
        cell.number_format = "#,##0"

for cell in ws["D"][1:]:             # 게시일
    cell.number_format = "yyyy-mm-dd"

for cell in ws["I"][1:]:             # 댓글작성일
    cell.number_format = "yyyy-mm-dd hh:mm"


# -------------------------
# Column Width
# -------------------------
widths = {
    "A": 8,    # 번호
    "B": 20,   # 크리에이터
    "C": 12,   # 콘텐츠유형
    "D": 14,   # 게시일
    "E": 18,   # 포스트ID
    "F": 24,   # 댓글ID
    "G": 20,   # 댓글작성자ID
    "H": 20,   # 댓글작성자
    "I": 18,   # 댓글작성일
    "J": 12,   # 댓글좋아요
    "K": 10,   # 댓글길이
    "L": 80,   # 댓글내용
    "M": 45,   # 포스트URL
}

for col, width in widths.items():
    ws.column_dimensions[col].width = width


# -------------------------
# Freeze / Filter
# -------------------------
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions


wb.save(filename)


# ------------------------------------------------
# 요약
# ------------------------------------------------
print(f"완료 : {filename}")
print(f"댓글 {len(df):,}개")

if "포스트ID" in df.columns:
    print(f"게시물 {df['포스트ID'].nunique():,}개")
if "댓글작성자ID" in df.columns:
    print(f"고유 작성자 {df['댓글작성자ID'].nunique():,}명")
if "콘텐츠유형" in df.columns:
    print("\n[유형별 댓글]")
    for t, n in df["콘텐츠유형"].value_counts().items():
        print(f"  {t:6s} {n:6,d}")