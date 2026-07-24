import os
import warnings
warnings.filterwarnings("ignore", category=UserWarning)
import pandas as pd
import pymysql

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from instagram.config import DB, OUTPUT_DIR
except Exception:
    from config import DB, OUTPUT_DIR

EXPORT_DIR = OUTPUT_DIR


# =====================================
# SQL — channel_metrics에서 계산된 값 그대로 export
# =====================================
sql = """
SELECT
    ROW_NUMBER() OVER (ORDER BY cr.nickname) AS 번호,
    cr.nickname AS 크리에이터,
    COALESCE(cr.agency,'-') AS 소속,

    MAX(chs.follower_count) AS follower_count,

    -- 통합 지표
    MAX(m.view_per_follower_ratio) AS vpf,
    MAX(m.engagement_rate)         AS er,
    MAX(m.loyalty_score)           AS loyalty,
    MAX(m.upload_frequency_weekly) AS upload_freq,
    MAX(m.avg_view)                AS total_avg_view,

    -- 표본 수 (해석 근거)
    MAX(m.sample_content_count) AS sample_cnt,
    MAX(m.aggregation_method)   AS agg_method,

    -- 최근 3개월
    MAX(m.videos_3m)          AS videos_3m,
    MAX(m.avg_view_3m)        AS avg_view_3m,
    MAX(m.avg_like_3m)        AS avg_like_3m,
    MAX(m.avg_comment_3m)     AS avg_comment_3m,
    MAX(m.engagement_rate_3m) AS engagement_rate_3m,

    -- Feed (DB는 longform 컬럼 사용)
    MAX(m.longform_avg_view)    AS longform_avg_view,
    MAX(m.longform_avg_like)    AS longform_avg_like,
    MAX(m.longform_avg_comment) AS longform_avg_comment,
    MAX(m.longform_er)          AS longform_er,
    MAX(m.longform_sample)      AS longform_sample,

    -- Reels (DB는 shorts 컬럼 사용)
    MAX(m.shorts_avg_view)    AS shorts_avg_view,
    MAX(m.shorts_avg_like)    AS shorts_avg_like,
    MAX(m.shorts_avg_comment) AS shorts_avg_comment,
    MAX(m.shorts_er)          AS shorts_er,
    MAX(m.shorts_sample)      AS shorts_sample,

    -- 광고 Feed
    MAX(m.ad_longform_avg_view)    AS ad_longform_avg_view,
    MAX(m.ad_longform_avg_like)    AS ad_longform_avg_like,
    MAX(m.ad_longform_avg_comment) AS ad_longform_avg_comment,
    MAX(m.ad_longform_er)          AS ad_longform_er,

    -- 일반 Feed
    MAX(m.normal_longform_avg_view)    AS normal_longform_avg_view,
    MAX(m.normal_longform_avg_like)    AS normal_longform_avg_like,
    MAX(m.normal_longform_avg_comment) AS normal_longform_avg_comment,
    MAX(m.normal_longform_er)          AS normal_longform_er,

    -- 광고 Reels
    MAX(m.ad_shorts_avg_view)    AS ad_shorts_avg_view,
    MAX(m.ad_shorts_avg_like)    AS ad_shorts_avg_like,
    MAX(m.ad_shorts_avg_comment) AS ad_shorts_avg_comment,
    MAX(m.ad_shorts_er)          AS ad_shorts_er,

    -- 일반 Reels
    MAX(m.normal_shorts_avg_view)    AS normal_shorts_avg_view,
    MAX(m.normal_shorts_avg_like)    AS normal_shorts_avg_like,
    MAX(m.normal_shorts_avg_comment) AS normal_shorts_avg_comment,
    MAX(m.normal_shorts_er)          AS normal_shorts_er,

    -- 댓글 지표
    MAX(m.commenter_overlap_rate)  AS commenter_overlap_rate,
    MAX(m.regular_commenter_count) AS regular_commenter_count,
    MAX(m.avg_comment_length)      AS avg_comment_length,
    MAX(m.l3_content_count)        AS l3_content_count

FROM creators cr

JOIN channels ch
    ON cr.creator_id = ch.creator_id

LEFT JOIN channel_snapshots chs
    ON ch.channel_id = chs.channel_id
   AND chs.captured_at = (
        SELECT MAX(captured_at)
        FROM channel_snapshots
        WHERE channel_id = ch.channel_id
   )

LEFT JOIN channel_metrics m
    ON ch.channel_id = m.channel_id

WHERE ch.platform='instagram'

GROUP BY ch.channel_id

ORDER BY cr.nickname;
"""

# =====================================
# SQL → DataFrame
# =====================================
conn = pymysql.connect(**DB)
try:
    df = pd.read_sql(sql, conn)
finally:
    conn.close()

print(f"조회 {len(df)}행")

# =====================================
# 컬럼 매핑 (계산 없이 channel_metrics 값 그대로)
# =====================================

df["팔로워"] = df["follower_count"]

df["조회수/팔로워(%)"] = df["vpf"]
df["공개참여율(ER)"] = df["er"]
df["업로드빈도(주)"] = df["upload_freq"]
df["Loyalty Score"] = df["loyalty"]
df["표본수"] = df["sample_cnt"]

df["최근3개월게시물"] = df["videos_3m"]
df["최근3개월평균조회수"] = df["avg_view_3m"]
df["최근3개월평균좋아요"] = df["avg_like_3m"]
df["최근3개월평균댓글"] = df["avg_comment_3m"]
df["최근3개월ER"] = df["engagement_rate_3m"]

df["전체평균조회수"] = df["total_avg_view"]

# ==========================
# Feed (longform 컬럼 사용)
# ==========================

df["피드평균조회수"] = df["longform_avg_view"]
df["피드평균좋아요"] = df["longform_avg_like"]
df["피드평균댓글"] = df["longform_avg_comment"]
df["피드ER"] = df["longform_er"]
df["피드표본"] = df["longform_sample"]

# ==========================
# Reels (shorts 컬럼 사용)
# ==========================

df["릴스평균조회수"] = df["shorts_avg_view"]
df["릴스평균좋아요"] = df["shorts_avg_like"]
df["릴스평균댓글"] = df["shorts_avg_comment"]
df["릴스ER"] = df["shorts_er"]
df["릴스표본"] = df["shorts_sample"]

# ==========================
# 광고 Feed
# ==========================

df["광고피드평균조회수"] = df["ad_longform_avg_view"]
df["광고피드평균좋아요"] = df["ad_longform_avg_like"]
df["광고피드평균댓글"] = df["ad_longform_avg_comment"]
df["광고피드ER"] = df["ad_longform_er"]

# ==========================
# 일반 Feed
# ==========================

df["일반피드평균조회수"] = df["normal_longform_avg_view"]
df["일반피드평균좋아요"] = df["normal_longform_avg_like"]
df["일반피드평균댓글"] = df["normal_longform_avg_comment"]
df["일반피드ER"] = df["normal_longform_er"]

# ==========================
# 광고 Reels
# ==========================

df["광고릴스평균조회수"] = df["ad_shorts_avg_view"]
df["광고릴스평균좋아요"] = df["ad_shorts_avg_like"]
df["광고릴스평균댓글"] = df["ad_shorts_avg_comment"]
df["광고릴스ER"] = df["ad_shorts_er"]

# ==========================
# 일반 Reels
# ==========================

df["일반릴스평균조회수"] = df["normal_shorts_avg_view"]
df["일반릴스평균좋아요"] = df["normal_shorts_avg_like"]
df["일반릴스평균댓글"] = df["normal_shorts_avg_comment"]
df["일반릴스ER"] = df["normal_shorts_er"]

# ==========================
# 댓글 지표
# ==========================

df["댓글작성자중복률(%)"] = df["commenter_overlap_rate"]
df["고정댓글러"] = df["regular_commenter_count"]
df["평균댓글길이"] = df["avg_comment_length"]
df["댓글수집콘텐츠"] = df["l3_content_count"]

# =====================================
# 최종 컬럼 순서
# =====================================
FINAL_COLS = [
    "번호",
    "크리에이터",
    "소속",
    "팔로워",

    "조회수/팔로워(%)",
    "공개참여율(ER)",
    "업로드빈도(주)",
    "Loyalty Score",
    "표본수",

    "최근3개월게시물",
    "최근3개월평균조회수",
    "최근3개월평균좋아요",
    "최근3개월평균댓글",
    "최근3개월ER",

    "전체평균조회수",

    "피드평균조회수",
    "피드평균좋아요",
    "피드평균댓글",
    "피드ER",
    "피드표본",

    "릴스평균조회수",
    "릴스평균좋아요",
    "릴스평균댓글",
    "릴스ER",
    "릴스표본",

    "광고피드평균조회수",
    "광고피드평균좋아요",
    "광고피드평균댓글",
    "광고피드ER",

    "일반피드평균조회수",
    "일반피드평균좋아요",
    "일반피드평균댓글",
    "일반피드ER",

    "광고릴스평균조회수",
    "광고릴스평균좋아요",
    "광고릴스평균댓글",
    "광고릴스ER",

    "일반릴스평균조회수",
    "일반릴스평균좋아요",
    "일반릴스평균댓글",
    "일반릴스ER",

    "댓글작성자중복률(%)",
    "고정댓글러",
    "평균댓글길이",
    "댓글수집콘텐츠",
]

# =====================================
# NaN → N/A
# =====================================
for col in FINAL_COLS:
    df[col] = df[col].apply(lambda x: "N/A" if pd.isna(x) else x)

df = df[FINAL_COLS]

# =====================================
# Excel 저장
# =====================================
os.makedirs(EXPORT_DIR, exist_ok=True)
filename = os.path.join(EXPORT_DIR, "instagram_파생지표.xlsx")
df.to_excel(filename, index=False, sheet_name="Metrics")

wb = load_workbook(filename)
ws = wb["Metrics"]

# =====================================
# 스타일
# =====================================
header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF")
center = Alignment(horizontal="center", vertical="center")
thin = Side(border_style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = center

# =====================================
# 숫자 서식 (헤더명 기반 — 순서 바뀌어도 안 깨짐)
# =====================================
fmt_map = {
    "팔로워":"#,##0",

    "조회수/팔로워(%)":"0.00",
    "공개참여율(ER)":"0.00",
    "업로드빈도(주)":"0.00",
    "Loyalty Score":"0.0000",
    "표본수":"#,##0",

    "최근3개월게시물":"#,##0",
    "최근3개월평균조회수":"#,##0.00",
    "최근3개월평균좋아요":"#,##0.00",
    "최근3개월평균댓글":"#,##0.00",
    "최근3개월ER":"0.00",

    "전체평균조회수":"#,##0.00",

    "피드평균조회수":"#,##0.00",
    "피드평균좋아요":"#,##0.00",
    "피드평균댓글":"#,##0.00",
    "피드ER":"0.00",
    "피드표본":"#,##0",

    "릴스평균조회수":"#,##0.00",
    "릴스평균좋아요":"#,##0.00",
    "릴스평균댓글":"#,##0.00",
    "릴스ER":"0.00",
    "릴스표본":"#,##0",

    "광고피드평균조회수":"#,##0.00",
    "광고피드평균좋아요":"#,##0.00",
    "광고피드평균댓글":"#,##0.00",
    "광고피드ER":"0.00",

    "일반피드평균조회수":"#,##0.00",
    "일반피드평균좋아요":"#,##0.00",
    "일반피드평균댓글":"#,##0.00",
    "일반피드ER":"0.00",

    "광고릴스평균조회수":"#,##0.00",
    "광고릴스평균좋아요":"#,##0.00",
    "광고릴스평균댓글":"#,##0.00",
    "광고릴스ER":"0.00",

    "일반릴스평균조회수":"#,##0.00",
    "일반릴스평균좋아요":"#,##0.00",
    "일반릴스평균댓글":"#,##0.00",
    "일반릴스ER":"0.00",

    "댓글작성자중복률(%)":"0.00",
    "고정댓글러":"#,##0",
    "평균댓글길이":"0.00",
    "댓글수집콘텐츠":"#,##0",
}
header_to_col = {}
for idx, cell in enumerate(ws[1], start=1):
    header_to_col[cell.value] = get_column_letter(idx)

for header, fmt in fmt_map.items():
    col_letter = header_to_col.get(header)
    if not col_letter:
        continue
    for cell in ws[col_letter][1:]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = fmt

# =====================================
# 열 너비 (헤더명 기반)
# =====================================
width_map = {
    "번호":8,
    "크리에이터":22,
    "소속":18,
    "팔로워":14,

    "조회수/팔로워(%)":16,
    "공개참여율(ER)":14,
    "업로드빈도(주)":14,
    "Loyalty Score":14,
    "표본수":10,

    "최근3개월게시물":16,
    "최근3개월평균조회수":18,
    "최근3개월평균좋아요":18,
    "최근3개월평균댓글":18,
    "최근3개월ER":14,

    "전체평균조회수":16,

    "피드평균조회수":16,
    "피드평균좋아요":16,
    "피드평균댓글":16,
    "피드ER":12,
    "피드표본":10,

    "릴스평균조회수":16,
    "릴스평균좋아요":16,
    "릴스평균댓글":16,
    "릴스ER":12,
    "릴스표본":10,

    "광고피드평균조회수":18,
    "광고피드평균좋아요":18,
    "광고피드평균댓글":18,
    "광고피드ER":14,

    "일반피드평균조회수":18,
    "일반피드평균좋아요":18,
    "일반피드평균댓글":18,
    "일반피드ER":14,

    "광고릴스평균조회수":18,
    "광고릴스평균좋아요":18,
    "광고릴스평균댓글":18,
    "광고릴스ER":14,

    "일반릴스평균조회수":18,
    "일반릴스평균좋아요":18,
    "일반릴스평균댓글":18,
    "일반릴스ER":14,

    "댓글작성자중복률(%)":18,
    "고정댓글러":12,
    "평균댓글길이":14,
    "댓글수집콘텐츠":14,
}
for header, width in width_map.items():
    col_letter = header_to_col.get(header)
    if col_letter:
        ws.column_dimensions[col_letter].width = width

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(filename)

print("=" * 60)
print(f"완료 : {filename}")
print(f"채널 수 : {len(df)}")
print("ℹ️ ER/Loyalty 는 팔로워 기준 (인스타는 피드 조회수를 제공하지 않음)")
print("ℹ️ 조회수 관련 지표는 릴스에만 값이 있음")
print("=" * 60)