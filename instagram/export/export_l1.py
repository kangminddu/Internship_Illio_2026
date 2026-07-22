# -*- coding: utf-8 -*-

import json
from pathlib import Path

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =========================================================
# 경로
# =========================================================
EXPORT_DIR = Path(__file__).resolve().parent
INSTA_DIR = EXPORT_DIR.parent
OUTPUT_DIR = INSTA_DIR / "output"

INPUT_FILE = OUTPUT_DIR / "l1_results.jsonl"
OUTPUT_FILE = OUTPUT_DIR / "instagram_l1.xlsx"

MAX_WIDTH = 40   # 컬럼 폭 상한

# 수집결과별 배경색 (검수 편의)
STATUS_COLORS = {
    "SUCCESS":        "C6EFCE",  # 초록
    "PRIVATE":        "FFEB9C",  # 노랑
    "NOT_FOUND":      "F2F2F2",  # 회색
    "LOGIN_REQUIRED": "FFC7CE",  # 빨강
    "CHALLENGE":      "FFC7CE",
    "RATE_LIMITED":   "FCE4D6",  # 주황
    "NETWORK_ERROR":  "FFC7CE",
    "TIMEOUT":        "FCE4D6",
    "ERROR":          "FFC7CE",
}


# =========================================================
# 엑셀 수식 오인 방지
# =========================================================
def sanitize_formula(value):
    """엑셀이 = + - @ 로 시작하는 문자열을 수식으로 인식하는 것을 방지."""
    if isinstance(value, str) and value:
        if value[0] in ("=", "+", "-", "@"):
            return "'" + value
    return value


# =========================================================
# JSONL 읽기 (append-only 로그)
# =========================================================
if not INPUT_FILE.exists():
    raise FileNotFoundError(f"결과 파일이 없습니다: {INPUT_FILE}")

rows = []
with INPUT_FILE.open("r", encoding="utf-8") as f:
    for line in f:
        line = line.strip()
        if not line:
            continue
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError:
            continue  # 깨진 줄은 건너뜀

if not rows:
    raise ValueError(f"읽을 레코드가 없습니다: {INPUT_FILE}")


# =========================================================
# dedup: username 기준 '최신' 레코드만 남김
#   JSONL은 append-only 이므로 '뒤에 온 줄'이 최신.
#   빈 username(None/""/공백)은 키가 뭉개지므로 제외.
# =========================================================
_raw_count = len(rows)

latest = {}
skipped_no_username = 0
for r in rows:
    uname = r.get("username")
    if not uname or not str(uname).strip():
        skipped_no_username += 1
        continue
    latest[uname] = r          # 뒤 줄이 앞 줄을 덮어씀 = 최신 유지

rows = list(latest.values())
_dedup_count = len(rows)

df = pd.DataFrame(rows)


# =========================================================
# 컬럼 순서 (JSONL에 없는 컬럼은 빈 값으로 생성 -> KeyError 방지)
# =========================================================
COLUMNS = [
    "username",
    "nickname", "user_id",
    "followers", "following", "posts",
    "account_type", "category_name",
    "is_private", "is_verified",
    "biography", "external_url",
    "status", "reason",
    "http_status",  "final_url",
    "profile_pic_url",
    "ts",
]

for col in COLUMNS:
    if col not in df.columns:
        df[col] = None        # 누락 컬럼 방어

df = df[COLUMNS]

ACCOUNT_TYPE_MAP = {
    1: "일반",
    2: "크리에이터",
    3: "비즈니스",
}

df["account_type"] = (
    pd.to_numeric(df["account_type"], errors="coerce")
    .map(ACCOUNT_TYPE_MAP)
    .fillna(df["account_type"])
)

# =========================================================
# 숫자 컬럼 명시적 숫자화 (문자열/None 섞여도 엑셀에서 숫자로 정렬되도록)
#   실패값은 NaN -> 엑셀 빈 칸
# =========================================================
for col in ["followers", "following", "posts"]:
    df[col] = pd.to_numeric(df[col], errors="coerce")


# =========================================================
# 컬럼명
# =========================================================
df.columns = [
    "아이디",
    "닉네임", "User ID",
    "팔로워", "팔로잉", "게시물",
    "계정유형", "카테고리",
    "비공개", "인증",
    "소개", "외부링크",
    "수집결과", "사유",
    "HTTP", "최종URL",
    "프로필사진",
    "수집시간",
]


# =========================================================
# 수식 오인 방지
# =========================================================
for col in ["아이디", "닉네임", "소개", "외부링크", "최종URL"]:
    df[col] = df[col].apply(sanitize_formula)


# =========================================================
# Excel 저장
# =========================================================
df.to_excel(OUTPUT_FILE, index=False, sheet_name="L1")


# =========================================================
# Excel 꾸미기
# =========================================================
wb = load_workbook(OUTPUT_FILE)
ws = wb["L1"]

header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF")
center = Alignment(horizontal="center", vertical="center")
thin = Side(border_style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 헤더명 -> 실제 엑셀 컬럼 문자 매핑 (컬럼 순서 바뀌어도 안전)
header_to_letter = {
    cell.value: get_column_letter(cell.column)
    for cell in ws[1]
}

# 헤더
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

# 데이터 셀 테두리
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border

# 숫자 포맷: 팔로워/팔로잉/게시물 (컬럼명 기준으로 위치 역산)
for name in ["팔로워", "팔로잉", "게시물"]:
    letter = header_to_letter.get(name)
    if not letter:
        continue
    for cell in ws[letter][1:]:
        cell.number_format = "#,##0"
        cell.alignment = center

# 수집결과 상태별 배경색 (실패가 눈에 띄도록)
result_letter = header_to_letter.get("수집결과")
if result_letter:
    for cell in ws[result_letter][1:]:
        color = STATUS_COLORS.get(cell.value)
        if color:
            cell.fill = PatternFill(fill_type="solid", fgColor=color)

# 컬럼 폭 자동 (상한 도달 시 조기 종료 -> 대량 데이터에서 순회 절약)
for column_cells in ws.columns:
    letter = get_column_letter(column_cells[0].column)
    length = 0
    for cell in column_cells:
        if cell.value is not None:
            length = max(length, len(str(cell.value)))
            if length >= MAX_WIDTH:
                break
    ws.column_dimensions[letter].width = min(length + 4, MAX_WIDTH)

# 고정/필터
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(OUTPUT_FILE)


# =========================================================
# 출력
# =========================================================
print(f"완료 : {OUTPUT_FILE}")
print(f"로그 {_raw_count}줄 -> dedup 후 {_dedup_count}개 계정")
if skipped_no_username:
    print(f"username 없음으로 제외 : {skipped_no_username}줄")
print("\n수집결과")
print(df["수집결과"].value_counts().to_string())