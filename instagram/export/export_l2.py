import os
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

import pandas as pd
import pymysql

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from instagram.config import DB, OUTPUT_DIR, L3_MIN_COMMENTS, L3_MAX_AGE_MONTHS


# ⚠️ pymysql 은 % 를 파라미터 플레이스홀더로 해석한다.
#    DATE_FORMAT('%Y-%m-%d') 같은 리터럴 % 는 %% 로 이스케이프해야
#    나중에 WHERE 조건(파라미터)을 추가해도 터지지 않는다.
#    여기서는 아예 DATE() 를 써서 % 자체를 없앴다.
sql = """
SELECT

ROW_NUMBER() OVER(
ORDER BY
cr.nickname,
ct.published_at DESC
) AS '번호',

cr.nickname AS '크리에이터',

COALESCE(cr.agency,'-') AS '소속',

DATE(ct.published_at) AS '게시일',

DATEDIFF(
CURDATE(),
ct.published_at
) AS '업로드경과일',

-- 수집 범위 구분. L2 는 최신 20개를 통째로 가져오므로
-- 1년 넘은 게시물도 섞여 있다. 분석 시 필터로 쓰라고 남긴다.
CASE
    WHEN ct.published_at > DATE_SUB(NOW(), INTERVAL 3 MONTH) THEN '3개월내'
    WHEN ct.published_at > DATE_SUB(NOW(), INTERVAL 6 MONTH) THEN '3-6개월'
    WHEN ct.published_at > DATE_SUB(NOW(), INTERVAL 12 MONTH) THEN '6-12개월'
    ELSE '1년+'
END AS '수집구간',

CASE
    WHEN ct.content_type='feed_image' THEN '피드'
    WHEN ct.content_type='carousel' THEN '캐러셀'
    WHEN ct.content_type='reels' THEN '릴스'
    ELSE ct.content_type
END AS '콘텐츠유형',

-- ⚠️ 릴스는 play_count(재생수), 피드/캐러셀은 view_count(노출).
--    피드/캐러셀은 인스타가 조회수를 아예 제공하지 않아 항상 NULL.
cs.view_count AS '조회수',

cs.like_count AS '좋아요',

cs.comment_count AS '댓글수',

CASE
    WHEN ct.is_paid_promotion=1 THEN 'O'
    ELSE 'X'
END AS '유료광고',

-- L3(댓글) 수집 대상 여부. 댓글 0개거나 1년 넘은 건 제외했다.
-- "이 게시물은 왜 댓글 데이터가 없냐"에 대한 답이 된다.
CASE
    WHEN ct.published_at > DATE_SUB(NOW(), INTERVAL %s MONTH)
         AND COALESCE(cs.comment_count,0) >= %s THEN 'O'
    ELSE '-'
END AS 'L3대상',

-- ⚠️ 릴스 캡션은 L2 릴스 탭 응답에 없어 상당수가 NULL.
--    L3 가 게시물 개별 방문 시 보강한다.
ct.caption_text AS '캡션',

-- 릴스는 /p/ 로 열면 /reels/ 로 리다이렉트되므로 유형별로 분기
CONCAT(
'https://www.instagram.com/',
CASE WHEN ct.content_type='reels' THEN 'reel/' ELSE 'p/' END,
ct.external_id,
'/'
) AS '포스트URL'

FROM creators cr

JOIN channels ch
ON cr.creator_id = ch.creator_id

JOIN contents ct
ON ch.channel_id = ct.channel_id

-- 콘텐츠별 최신 스냅샷 1건만
LEFT JOIN (

SELECT cs1.*

FROM content_snapshots cs1

JOIN (

SELECT
content_id,
MAX(captured_at) latest

FROM content_snapshots

GROUP BY content_id

) t

ON cs1.content_id=t.content_id
AND cs1.captured_at=t.latest

) cs

ON ct.content_id=cs.content_id

WHERE ch.platform='instagram'

ORDER BY
cr.nickname,
ct.published_at DESC;
"""

PARAMS = [L3_MAX_AGE_MONTHS, L3_MIN_COMMENTS]


def sanitize_formula(val):
    """엑셀이 수식으로 인식하는 것 방지"""
    if isinstance(val, str) and val and val[0] in ("=", "+", "-", "@"):
        return "'" + val
    return val


# -------------------------------------------------
# DB → DataFrame
# -------------------------------------------------

conn = pymysql.connect(**DB)
try:
    df = pd.read_sql(sql, conn, params=PARAMS)
finally:
    conn.close()

for col in ["크리에이터", "소속", "캡션"]:
    if col in df.columns:
        df[col] = df[col].fillna("").apply(sanitize_formula)


os.makedirs(OUTPUT_DIR, exist_ok=True)
filename = os.path.join(OUTPUT_DIR, "L2_인스타그램_게시물.xlsx")
df.to_excel(filename, index=False, sheet_name="L2")


# ==================================================
# Excel Styling
# ==================================================

wb = load_workbook(filename)
ws = wb["L2"]

header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF")

center = Alignment(horizontal="center", vertical="center")
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=False)

thin = Side(border_style="thin", color="D9D9D9")
border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)

# -------------------------
# Header
# -------------------------

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

# -------------------------
# Body
# -------------------------

# 가운데 정렬: 번호/게시일/경과일/수집구간/유형/조회수/좋아요/댓글수/유료광고/L3대상
# (캡션 M, 포스트URL N 만 왼쪽 정렬)
CENTER_COLS = {1, 4, 5, 6, 7, 8, 9, 10, 11, 12}

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = center if cell.column in CENTER_COLS else left_wrap

# -------------------------
# Number / Date Format
# -------------------------

for col in ["H", "I", "J"]:          # 조회수 / 좋아요 / 댓글수
    for cell in ws[col][1:]:
        cell.number_format = "#,##0"

for cell in ws["D"][1:]:             # 게시일
    cell.number_format = "yyyy-mm-dd"

# -------------------------
# Width
# -------------------------

widths = {
    "A": 8,    # 번호
    "B": 20,   # 크리에이터
    "C": 16,   # 소속
    "D": 14,   # 게시일
    "E": 14,   # 업로드경과일
    "F": 12,   # 수집구간
    "G": 12,   # 콘텐츠유형
    "H": 12,   # 조회수
    "I": 12,   # 좋아요
    "J": 12,   # 댓글수
    "K": 10,   # 유료광고
    "L": 10,   # L3대상
    "M": 70,   # 캡션
    "N": 45,   # 포스트URL
}

for col, width in widths.items():
    ws.column_dimensions[col].width = width

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

wb.save(filename)


# ==================================================
# 요약 출력 — 데이터 상태를 눈으로 확인
# ==================================================

print(f"완료 : {filename}")
print(f"게시물 {len(df):,}개")

if "콘텐츠유형" in df.columns:
    print("\n[유형별]")
    for t, n in df["콘텐츠유형"].value_counts().items():
        print(f"  {t:6s} {n:6,d}")

if "수집구간" in df.columns:
    print("\n[기간별]")
    order = ["3개월내", "3-6개월", "6-12개월", "1년+"]
    vc = df["수집구간"].value_counts()
    for k in order:
        if k in vc:
            print(f"  {k:8s} {vc[k]:6,d}")

if "L3대상" in df.columns:
    n_t = (df["L3대상"] == "O").sum()
    print(f"\nL3(댓글) 수집 대상 {n_t:,}개 "
          f"(댓글 {L3_MIN_COMMENTS}개 이상 + {L3_MAX_AGE_MONTHS}개월 이내)")

if "캡션" in df.columns:
    empty_cap = (df["캡션"] == "").sum()
    if empty_cap:
        print(f"⚠️ 캡션 없음 {empty_cap:,}개 "
              f"({empty_cap / len(df) * 100:.0f}%) — 대부분 릴스, L3 에서 보강 예정")

for c in ("좋아요", "댓글수"):
    if c in df.columns:
        miss = df[c].isna().sum()
        if miss:
            print(f"⚠️ {c} 결측 {miss:,}개 (스냅샷 누락)")

# 조회수는 릴스에만 존재하는 지표 — 피드/캐러셀 NULL 은 정상
if {"조회수", "콘텐츠유형"} <= set(df.columns):
    reels_miss = df[(df["콘텐츠유형"] == "릴스") & df["조회수"].isna()].shape[0]
    if reels_miss:
        print(f"⚠️ 릴스 조회수 결측 {reels_miss:,}개")
    print("ℹ️ 피드/캐러셀은 인스타가 조회수를 제공하지 않음 (정상)")