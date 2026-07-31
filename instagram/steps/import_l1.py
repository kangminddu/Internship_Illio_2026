# -*- coding: utf-8 -*-
"""
steps/import_l1.py  (v3 — 시트 기반 key 재매핑)

L1 결과(output/l1_results.jsonl) → fandom_crm.channels 적재.

★ 이 단계는 인스타에만 있다.
------
    유튜브 : crawler_l1_parallel이 크롤하면서 바로 DB에 쓴다
    틱톡   : steps/l1이 크롤하면서 바로 DB에 쓴다
    인스타 : steps/l1이 jsonl 파일에 쓰고 → import_l1이 DB에 적재  ← 2단계

왜 나눴나:
  ① 인스타 L1은 계정당 8~15초라 1,881개에 4~8시간이 걸린다.
     크롤 도중 DB 스키마나 적재 규칙을 바꾸고 싶어도
     다시 크롤할 수 없다. → 원본을 파일로 남기고 적재만 다시 돌린다.
  ② 크롤러가 DB를 모르면 세션 관리에만 집중할 수 있다.
     (Playwright 단일 세션을 8시간 유지하는 게 이 단계의 핵심이다)
  ③ jsonl은 한 줄씩 append되므로 중간에 죽어도 앞부분이 남는다.

■ 적재 규칙
  SUCCESS   → channel_id_status='resolved',  existence='normal'   ← L2 대상
  PRIVATE   → 'resolved',   'private'
  NOT_FOUND → 'not_found',  'deleted'
  그 외(ERROR/TIMEOUT/CHALLENGE)는 미확정이라 적재하지 않음.

  ★ '미확정은 적재하지 않는다'가 핵심이다.
    TIMEOUT을 'deleted'로 적으면 멀쩡한 계정이 영구 제외된다.
    확정 실패(NOT_FOUND)와 불확실한 실패(ERROR/TIMEOUT)를 구분하는
    유튜브 classify_existence()와 같은 원칙.
    적재하지 않으면 channel_id_status가 'handle_only'로 남아
    다음 L1 실행에서 다시 대상이 된다.

  PRIVATE을 'resolved'로 두는 이유: 계정은 존재하고 프로필 정보도
  얻었다. 다만 게시물을 못 볼 뿐이라 L2에서 걸러야 한다.

channel_activity_status 는 항상 'unknown'. 활동성 판단은 metrics 몫.
  ↑ 세 플랫폼 중 인스타만 이렇다.
    유튜브 : L2a 잠정 → backfill 확정
    틱톡   : L2에서 확정
    인스타 : calc_metric 안에서 판정
    → 판정 시점이 전부 다르다. 리뷰 안건.
"""

import json
from collections import Counter, defaultdict

import pymysql
from openpyxl import load_workbook

try:
    from instagram.config import DB, RESULTS_FILE
    from instagram.reader import DEFAULT_EXCEL, extract_username
except Exception:
    from config import DB, RESULTS_FILE
    from reader import DEFAULT_EXCEL, extract_username


# =========================================================
# 설정
# =========================================================
DRY_RUN = False          # True 면 DB 에 쓰지 않고 집계만 출력
                         # ⚠️ 상수라 CLI로 못 바꾼다. seed.py는 --dry-run
                         #    인자가 있는데 여기는 코드를 고쳐야 한다.
COMMIT_EVERY = 200

# L1 status → (channel_id_status, channel_existence_status)
# 두 컬럼이 다른 질문에 답한다:
#   id_status        "이 URL로 계정을 특정할 수 있나"
#   existence_status "계정이 살아있나"
STATUS_MAP = {
    "SUCCESS":   ("resolved",  "normal"),
    "PRIVATE":   ("resolved",  "private"),
    "NOT_FOUND": ("not_found", "deleted"),
}

# 같은 계정이 여러 번 크롤됐을 때(재실행 등) 어느 결과를 쓸지.
# 확정도가 높은 쪽이 이긴다.
#   SUCCESS(3) > PRIVATE(2) > NOT_FOUND(1) > 그 외(0)
# 한 번이라도 SUCCESS면 그걸 쓴다 — 일시적 실패로 덮이지 않게.
STATUS_PRIORITY = {"SUCCESS": 3, "PRIVATE": 2, "NOT_FOUND": 1}

# IG account_type(숫자) → channels.account_type(varchar)
# TODO: 실제 데이터로 매핑 검증 필요. 모르는 값은 숫자 문자열 그대로 저장.
#
# ★ 모르는 값을 버리지 않고 숫자 문자열로 남긴다.
#   나중에 DB를 조회해서 "4가 뭐지?"를 조사할 수 있다.
#   None으로 버리면 그런 값이 있었다는 사실 자체가 사라진다.
ACCOUNT_TYPE_MAP = {1: "일반", 2: "크리에이터", 3: "비즈니스"}


SQL_UPSERT_CHANNEL = """
INSERT INTO channels (
    creator_id, platform,
    channel_url_raw, channel_url_normalized,
    channel_id_status, channel_existence_status, channel_activity_status,
    external_channel_id, channel_name,
    account_type, bio, external_link
) VALUES (
    %s, 'instagram',
    %s, %s,
    %s, %s, 'unknown',
    %s, %s,
    %s, %s, %s
)
ON DUPLICATE KEY UPDATE
    creator_id=VALUES(creator_id),
    channel_url_raw=VALUES(channel_url_raw),
    channel_url_normalized=VALUES(channel_url_normalized),
    channel_id_status=VALUES(channel_id_status),
    channel_existence_status=VALUES(channel_existence_status),
    external_channel_id=VALUES(external_channel_id),
    channel_name=VALUES(channel_name),
    account_type=VALUES(account_type),
    bio=VALUES(bio),
    external_link=VALUES(external_link),
    updated_at=CURRENT_TIMESTAMP
"""
# ⚠️ activity_status는 UPDATE 목록에 없다.
#   재실행해도 metrics가 매긴 활동성이 'unknown'으로 되돌아가지 않는다.
#   (유튜브 L1이 channel_opened_at을 COALESCE로 보호하는 것과 같은 발상)

SQL_INSERT_CHANNEL_SNAPSHOT = """
INSERT INTO channel_snapshots (
    channel_id,
    captured_at,
    follower_count,
    following_count,
    total_video_count,
    total_like_count,
    total_view_count
)
VALUES (
    %s,
    NOW(),
    %s,
    %s,
    %s,
    NULL,
    NULL
)
ON DUPLICATE KEY UPDATE
    follower_count=VALUES(follower_count),
    following_count=VALUES(following_count),
    total_video_count=VALUES(total_video_count),
    total_like_count=VALUES(total_like_count),
    total_view_count=VALUES(total_view_count);
"""
# total_like_count / total_view_count가 NULL인 이유:
# 인스타는 계정 단위 누적 좋아요/조회수를 공개하지 않는다.
# 유튜브는 채널 총조회수를 주고, 틱톡은 heartCount를 준다.
# → 같은 스키마를 쓰지만 플랫폼마다 채울 수 있는 컬럼이 다르다.
#
# NOW()를 쓰는 것도 눈여겨볼 부분.
# 파이썬에서 datetime을 만들어 넘기면 pymysql이 tzinfo를 버려
# 시간대 사고가 났었다. MySQL이 직접 자기 시각을 쓰면 그 문제가 없다.


# =========================================================
# 헬퍼
# =========================================================
def normalize_url(username):
    """seed.py의 normalize_url과 반드시 동일해야 한다.
    이 URL로 channels를 조회하므로, 형식이 어긋나면 매칭이 전부 실패한다."""
    return f"https://www.instagram.com/{username}/"


def account_type_label(v):
    """숫자 → 한글 라벨. 모르는 값은 숫자 문자열 그대로."""
    if v is None:
        return None
    try:
        return ACCOUNT_TYPE_MAP.get(int(v), str(v))
    except (TypeError, ValueError):
        return str(v)


def is_better(new, old):
    """확정 상태 우선, 동순위면 최신 ts.

    같은 계정이 jsonl에 여러 번 나오는 경우:
      - L1을 여러 번 돌렸다
      - 시드 엑셀에 같은 계정이 여러 키값으로 들어있다

    한 번은 SUCCESS, 한 번은 TIMEOUT이면 SUCCESS를 써야 한다.
    시간순으로만 고르면 나중의 TIMEOUT이 이겨서 데이터를 잃는다.
    """
    if old is None:
        return True
    pn = STATUS_PRIORITY.get(new.get("status"), 0)
    po = STATUS_PRIORITY.get(old.get("status"), 0)
    if pn != po:
        return pn > po
    return (new.get("ts") or "") >= (old.get("ts") or "")
    # ↑ 같은 등급이면 최신. ISO 문자열이라 사전순 비교가 곧 시간순 비교다.


def load_jsonl():
    """jsonl 로드. 깨진 줄은 건너뛴다.

    한 줄이 깨졌다고 전체를 못 읽으면 8시간짜리 크롤 결과가 날아간다.
    → 예외를 던지지 않고 그 줄만 버린다.
    (크롤 중 강제 종료되면 마지막 줄이 잘려 있는 경우가 실제로 있다)
    """
    if not RESULTS_FILE.exists():
        raise FileNotFoundError(RESULTS_FILE)
    rows = []
    with RESULTS_FILE.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return rows


def build_username_key_map():
    """
    시트 전체(dedup 없이)를 읽어 username(소문자) → [key, ...] 매핑을 만든다.
    reader.py 가 버린 중복 행의 key 를 되살리기 위한 것.

    ★ v3에서 추가된 부분. 왜 필요했나:

      reader.py는 username 기준으로 dedup한다. 같은 계정이 여러 키값에
      물려 있으면 '첫 번째' 키값만 남기고 나머지를 버린다.

      그런데 그 첫 번째 키값이 creators 테이블에 없을 수 있다.
      (시드 적재 시점의 필터나 순서에 따라)
      그러면 계정은 크롤에 성공했는데 붙일 creator가 없어서 버려진다.

      → 시트를 다시 읽어 그 계정의 '모든' 키값 후보를 확보하고,
        하나씩 시도해서 creators에 있는 것을 찾는다.

      아래 main의 rescued 카운터가 이 방식으로 살린 건수다.
    """
    wb = load_workbook(DEFAULT_EXCEL, data_only=True)
    ws = wb["Sheet1"]
    m = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 3:
            continue
        key, insta = row[0], row[2]
        if not key or insta is None:
            continue
        raw = str(insta).strip()
        if not raw or raw.upper() == "NULL":
            continue
        u = extract_username(raw)   # reader.py와 동일 로직
        if not u:
            continue
        k = str(key).strip()
        if k not in m[u.lower()]:
            m[u.lower()].append(k)   # 순서 보존 (먼저 나온 키를 먼저 시도)
    wb.close()
    return m


def load_creator_map(conn):
    """seed_key → creator_id

    전체를 한 번에 메모리로 읽는다. 계정마다 SELECT를 날리면
    수천 번 왕복이 된다.
    """
    with conn.cursor() as cur:
        cur.execute("SELECT seed_key, creator_id FROM creators "
                    "WHERE seed_key IS NOT NULL")
        return {r[0]: r[1] for r in cur.fetchall()}


def dedup_by_username(rows):
    """같은 username의 여러 결과 중 최선 하나만 남긴다."""
    best = {}
    no_username = 0
    for r in rows:
        u = (r.get("username") or "").strip().strip("/")
        if not u:
            no_username += 1
            continue
        k = u.lower()   # 인스타 username은 대소문자 무관
        if is_better(r, best.get(k)):
            r["_username"] = u   # 정리된 username을 붙여둔다
            best[k] = r
    return best, no_username


# =========================================================
# 메인
# =========================================================
def main():
    rows = load_jsonl()
    best, no_username = dedup_by_username(rows)
    status_counts = Counter(r.get("status") for r in best.values())

    # 적재 전 전체 그림. 상태별로 '적재/제외'를 명시해서
    # 왜 이 건수가 안 들어가는지 바로 알 수 있게 한다.
    print("=" * 58)
    print(f"JSONL 총 라인      : {len(rows)}")
    print(f"유니크 username    : {len(best)}")
    if no_username:
        print(f"username 없음(스킵): {no_username}")
    for st, c in status_counts.most_common():
        mark = "적재" if st in STATUS_MAP else "제외"
        print(f"  {st:<16} {c:>6}  ({mark})")
    print("=" * 58)

    print("시트에서 username→key 후보 매핑 생성 중...")
    uname_keys = build_username_key_map()
    print(f"  시트 인스타 username: {len(uname_keys)}개")

    conn = pymysql.connect(**DB)
    try:
        creator_map = load_creator_map(conn)
        print(f"  creators 로드: {len(creator_map)}명")
        print("=" * 58)

        upserted = skipped = 0
        no_creator = 0
        rescued = 0           # jsonl key 로는 실패했는데 시트 후보 key 로 살린 건수
        rescued_ex, missing_ex = [], []
        # ↑ 예시를 5개만 모은다. 전부 출력하면 로그가 넘치고,
        #   0개면 무슨 일이 있었는지 알 수 없다.

        with conn.cursor() as cur:
            for i, row in enumerate(best.values(), 1):
                status = row.get("status")
                if status not in STATUS_MAP:
                    skipped += 1
                    continue      # ERROR/TIMEOUT/CHALLENGE → 미확정, 적재 안 함

                username = row["_username"]
                jkey = row.get("key")

                # 후보 key: jsonl 의 key 를 먼저, 그 다음 시트에서 나온 모든 key
                #
                # 순서가 중요하다. jsonl의 key가 원래 의도된 매핑이고,
                # 시트 후보는 그게 실패했을 때의 구제책이다.
                candidates = []
                if jkey:
                    candidates.append(str(jkey).strip())
                for k in uname_keys.get(username.lower(), []):
                    if k not in candidates:
                        candidates.append(k)

                creator_id = None
                used_key = None
                for k in candidates:
                    if k in creator_map:
                        creator_id = creator_map[k]
                        used_key = k
                        break      # 첫 번째로 찾은 것을 쓴다

                if creator_id is None:
                    # 후보를 다 시도했는데 creators에 없다.
                    # 시드가 안 돌았거나, 그 키값이 필터에서 빠진 경우.
                    no_creator += 1
                    if len(missing_ex) < 5:
                        missing_ex.append(f"@{username}({jkey})")
                    continue

                if used_key != (str(jkey).strip() if jkey else None):
                    # jsonl의 key가 아닌 다른 key로 찾았다 = 구제됨
                    rescued += 1
                    if len(rescued_ex) < 5:
                        rescued_ex.append(f"@{username} {jkey}→{used_key}")

                id_status, existence = STATUS_MAP[status]

                if not DRY_RUN:
                    cur.execute(SQL_UPSERT_CHANNEL, (
                        creator_id,
                        row.get("url") or normalize_url(username),
                        normalize_url(username),
                        id_status,
                        existence,
                        row.get("user_id"),        # 인스타 내부 ID.
                        #  ↑ username은 바뀔 수 있지만 user_id는 영구하다.
                        #    (유튜브 UC ID, 틱톡 sec_uid와 같은 역할)
                        row.get("nickname"),
                        account_type_label(row.get("account_type")),
                        row.get("biography"),
                        row.get("external_url"),
                    ))
                    # ON DUPLICATE 경로에서는 lastrowid를 신뢰할 수 없어
                    # channel_id를 직접 조회한다.
                    # ⚠️ 계정마다 SELECT가 하나 더 나간다.
                    #    LAST_INSERT_ID(channel_id) 트릭을 쓰면 없앨 수 있다.
                    #    (틱톡 l2.py의 INSERT_CONTENT가 그 방식)
                    cur.execute("""
                        SELECT channel_id
                        FROM channels
                        WHERE platform='instagram'
                        AND channel_url_normalized=%s
                    """, (normalize_url(username),))

                    channel_id = cur.fetchone()[0]
                    cur.execute(
                        SQL_INSERT_CHANNEL_SNAPSHOT,
                        (
                            channel_id,
                            row.get("followers"),
                            row.get("following"),
                            row.get("media_count"),
                        ),
                    )
                    if i % COMMIT_EVERY == 0:
                        conn.commit()
                upserted += 1

        if not DRY_RUN:
            conn.commit()
    finally:
        conn.close()
        # ⚠️ 예외 시 rollback이 없다. seed.py는 try/except로 롤백하는데
        #    여기는 finally에 close만 있다. 중간에 죽으면
        #    마지막 커밋 이후 분이 그대로 날아간다.
        #    (COMMIT_EVERY=200이라 손실이 작긴 하다)

    print(f"적재(upsert)   : {upserted}" + ("  [DRY_RUN, 미반영]" if DRY_RUN else ""))
    print(f"  └ 시트 key 로 복구: {rescued}")
    if rescued_ex:
        print(f"     예시: {', '.join(rescued_ex)}")
    print(f"건너뜀(미확정) : {skipped}")
    print(f"creator 없음   : {no_creator}")
    if missing_ex:
        print(f"     예시: {', '.join(missing_ex)}")
    print("=" * 58)


if __name__ == "__main__":
    main()