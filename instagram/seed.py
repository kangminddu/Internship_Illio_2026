# -*- coding: utf-8 -*-
"""
seed_instagram.py

SNS_정보.xlsx Sheet1 의 ChInstagram 컬럼 → creators + channels(platform='instagram') 적재.
crawler/seed.py 의 인스타 버전이지만, 결정적으로 다른 점이 하나 있다.

■ seed.py 를 그대로 쓰면 안 되는 이유
seed.py 는 seed_key 를 '행 번호'로 만든다:  seed_key = f"{prefix}_{i}"
그런데 Sheet1 의 '키값' 컬럼은 행 번호가 아니라 외부 마스터 ID 다.
  - Sheet1 은 23728행인데 키값 최대치는 G_236822
  - Sheet1 의 35번째 행의 실제 키값은 G_281
따라서 행 번호로 키를 만들면 G_2210(이미 존재하는 다른 크리에이터)에
엉뚱한 데이터가 붙어 기존 646명이 오염된다.
→ 이 스크립트는 '키값' 컬럼을 그대로 seed_key 로 쓴다.

★ 이게 이 파일의 핵심이다.
------
  유튜브 seed.py를 그대로 복사해 썼다면 기존 크리에이터 646명의
  데이터가 엉뚱한 계정으로 덮였을 것이다. 실행 전에 발견했다.

  왜 유튜브에서는 문제가 없었나:
    유튜브 시드는 자기 엑셀로 creators를 '새로 만드는' 쪽이라
    행 번호 키가 자체적으로 일관되면 됐다.
    인스타 시드는 '이미 있는 creator에 채널을 붙이는' 쪽이라
    기존 키와 정확히 맞아야 한다.

  → 같은 코드를 다른 맥락에 옮길 때 '전제 조건'까지 함께 오지 않는다.
    (backfill_activity.py에서 겪은 것과 같은 교훈)

■ 기존 creators 보호
이미 있는 seed_key 는 nickname/category/agency 를 건드리지 않는다.
(ON DUPLICATE KEY UPDATE 에서 creator_id 만 회수)
신규 생성 시 nickname 은 seed_key 를 임시값으로 넣는다.
  → 기존 관행과 동일. crawler_l1.py 가 유튜브에서
    "UPDATE creators SET nickname=... WHERE nickname LIKE 'G\\_%'" 로 덮어쓰는 패턴.
  → 인스타는 import_l1.py 가 L1 의 full_name 으로 덮어쓰면 된다.

■ username 추출은 reader.py 것을 그대로 재사용
seed 가 만드는 URL 과 L1 이 크롤한 URL 이 반드시 일치해야
import_l1 이 채널을 찾을 수 있다.

  ↑ 이 제약이 중요하다. import_l1은 channel_url_normalized로
    채널을 찾는데, seed가 만든 URL과 L1이 만든 URL의 형식이
    한 글자라도 다르면(끝 슬래시 유무 등) 매칭이 전부 실패한다.
    → normalize_url()을 seed/import_l1이 각각 갖고 있되
      내용이 동일해야 한다. (공용 모듈로 빼는 게 맞다)

사용법:
    python -m instagram.seed_instagram              # 실제 적재
    python -m instagram.seed_instagram --dry-run    # 집계만
"""

import argparse
from collections import defaultdict

import pymysql
from openpyxl import load_workbook

# try/except import: 패키지로 실행(-m instagram.seed)할 때와
# 디렉터리 안에서 직접 실행할 때 둘 다 되게 한 것.
# 개발 중 편의를 위한 장치인데, 예외를 통째로 삼켜서
# 진짜 import 에러(오타 등)도 폴백으로 넘어가 원인을 가린다.
try:
    from instagram.config import DB
    from instagram.reader import DEFAULT_EXCEL, extract_username
except Exception:
    from config import DB
    from reader import DEFAULT_EXCEL, extract_username


SHEET = "Sheet1"
KEY_COL = 0          # A: 키값
INSTA_COL = 2        # C: ChInstagram
# ⚠️ 헤더명이 아니라 인덱스로 하드코딩. 열이 하나 삽입되면 조용히 어긋난다.
#    유튜브 seed.py는 헤더명+별칭 사전으로 찾아서 순서 변경에 견딘다.


# 기존 creator 는 절대 건드리지 않고 creator_id 만 회수한다.
#
# ★ ON DUPLICATE KEY UPDATE creator_id=LAST_INSERT_ID(creator_id)
#   이 관용구가 두 가지를 동시에 한다:
#     ① 아무 컬럼도 실제로 바꾸지 않는다 (creator_id를 자기 자신으로)
#     ② lastrowid로 기존 creator_id를 받아올 수 있게 한다
#   그냥 "seed_key=seed_key"로 쓰면 ①은 되지만 ②가 안 돼서
#   별도 SELECT가 필요하다.
SQL_UPSERT_CREATOR = """
INSERT INTO creators (seed_key, nickname)
VALUES (%s, %s)
ON DUPLICATE KEY UPDATE creator_id=LAST_INSERT_ID(creator_id)
"""

SQL_UPSERT_CHANNEL = """
INSERT INTO channels
  (creator_id, platform, channel_url_raw, channel_url_normalized,
   channel_id_status, channel_existence_status, channel_activity_status, is_primary)
VALUES (%s, 'instagram', %s, %s, 'handle_only', 'unknown', 'unknown', %s)
ON DUPLICATE KEY UPDATE
  channel_url_normalized=VALUES(channel_url_normalized),
  updated_at=CURRENT_TIMESTAMP
"""
# 세 상태값을 모두 초기값으로 넣는다:
#   channel_id_status='handle_only'  → username은 있지만 user_id는 모름.
#                                       L1이 성공하면 import_l1이 'resolved'로 승격
#   existence='unknown'              → 계정이 살아있는지 모름. L1이 확정
#   activity='unknown'               → 활동성은 metrics 몫 (인스타 특유)


def normalize_url(username):
    """L1/import_l1 과 동일한 규칙이어야 한다.

    끝 슬래시가 있다. import_l1이 이 형식으로 채널을 찾으므로
    한쪽만 바꾸면 매칭이 전부 실패한다.
    """
    return f"https://www.instagram.com/{username}/"


def scan_sheet(excel_path):
    """
    Sheet1 스캔 → [(seed_key, username, raw_url), ...]
    (seed_key, username) 중복은 제거. 완전 동일 행이 12854개라 필수.

    ★ 23,728행 중 12,854행이 완전 중복이다.
      엑셀이 어떤 시스템에서 export되면서 행이 뻥튀기된 것으로 보인다.
      dedup 없이 돌리면 같은 INSERT를 12,854번 더 날린다.

    reader.py의 load_instagram_urls와 비슷하지만 dedup 기준이 다르다:
      reader   : username만 기준 (같은 계정은 한 번만)
      여기     : (seed_key, username) 쌍 기준
                 → 같은 계정을 여러 크리에이터가 가리키는 경우를 보존한다
                   (아래 main의 multi 집계에서 확인)
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    # read_only=True: 23,728행을 메모리에 다 올리지 않고 스트리밍으로 읽는다.
    # data_only=True: 수식 셀에서 계산된 값을 읽는다.
    ws = wb[SHEET]

    pairs = []
    seen = set()
    # 사유별 집계 = 시드 엑셀의 품질 리포트
    stats = {"rows": 0, "no_key": 0, "no_insta": 0, "bad_username": 0, "dup": 0}

    for row in ws.iter_rows(min_row=2, values_only=True):
        stats["rows"] += 1
        if len(row) <= INSTA_COL:
            continue    # 뒤쪽이 전부 빈 행은 openpyxl이 짧게 반환한다

        key = row[KEY_COL]
        insta = row[INSTA_COL]

        if not key or not str(key).strip():
            stats["no_key"] += 1
            continue
        seed_key = str(key).strip()

        if insta is None:
            stats["no_insta"] += 1
            continue
        raw = str(insta).strip()
        if not raw or raw.upper() == "NULL":
            stats["no_insta"] += 1
            continue

        username = extract_username(raw)      # reader.py 와 동일 로직
        # ↑ 복사가 아니라 import다. 두 곳에 같은 로직을 두면
        #   한쪽만 고쳐서 URL이 갈리고, import_l1의 매칭이 깨진다.
        if not username:
            stats["bad_username"] += 1
            continue

        sig = (seed_key, username.lower())
        if sig in seen:
            stats["dup"] += 1
            continue
        seen.add(sig)
        pairs.append((seed_key, username, raw))

    wb.close()
    return pairs, stats


def main(excel_path=None, dry_run=False):
    excel_path = excel_path or DEFAULT_EXCEL
    pairs, stats = scan_sheet(excel_path)

    usernames = {u.lower() for _, u, _ in pairs}
    keys = {k for k, _, _ in pairs}

    # 한 계정을 여러 creator 가 가리키는 경우 (채널은 1행만 생김 → 마지막이 이김)
    #
    # ★ 이걸 집계해서 경고로 보여준다.
    #   소속사 공식 계정을 소속 크리에이터 여러 명이 각자 적어둔 경우가 있다.
    #   channels.channel_url_normalized가 UNIQUE라 행은 하나만 생기고,
    #   마지막에 처리된 creator_id가 남는다.
    #   → 데이터가 조용히 틀리는 대신 "이런 일이 있다"를 드러낸다.
    owners = defaultdict(set)
    for k, u, _ in pairs:
        owners[u.lower()].add(k)
    multi = {u: ks for u, ks in owners.items() if len(ks) > 1}

    # 적재 전에 전체 그림을 먼저 출력한다.
    # 23,728행을 넣고 나서 "엉뚱한 걸 읽었다"를 발견하면 늦다.
    print("=" * 62)
    print(f"엑셀: {excel_path}")
    print(f"전체 행                 : {stats['rows']}")
    print(f"  키값 없음             : {stats['no_key']}")
    print(f"  인스타 없음/NULL      : {stats['no_insta']}")
    print(f"  username 추출 실패    : {stats['bad_username']}")
    print(f"  (키값,계정) 중복 제거 : {stats['dup']}")
    print("-" * 62)
    print(f"적재 대상 (키값,계정) 쌍: {len(pairs)}")
    print(f"  유니크 seed_key       : {len(keys)}")
    print(f"  유니크 인스타 계정    : {len(usernames)}")
    if multi:
        print(f"  ⚠️ 한 계정을 여러 키값이 가리킴: {len(multi)}건 (채널 1행만 생성됨)")
        for u, ks in list(multi.items())[:5]:   # 예시 5개만
            print(f"     @{u}: {sorted(ks)}")
    print("=" * 62)

    # ★ --dry-run: DB를 건드리기 전에 결과를 확인한다.
    #   seed는 수만 행을 한 번에 넣는 작업이라 되돌리기 어렵다.
    #   기존 creator 646명이 오염될 뻔한 사고를 겪은 뒤 넣은 안전장치.
    if dry_run:
        print("DRY-RUN → DB 미반영. 종료.")
        return

    conn = pymysql.connect(**DB, autocommit=False)   # 명시적 트랜잭션
    created_creator = existing_creator = 0
    ch_upserted = 0
    primary_done = set()

    try:
        with conn.cursor() as cur:
            # 기존 creators 파악 (신규 생성 수를 정확히 세기 위해)
            #
            # ON DUPLICATE는 "만들었는지 재사용했는지"를 알려주지 않는다.
            # 시작 전 스냅샷을 떠두고 비교해야 정확한 집계가 나온다.
            cur.execute("SELECT seed_key FROM creators WHERE seed_key IS NOT NULL")
            before = {r[0] for r in cur.fetchall()}

            for i, (seed_key, username, raw) in enumerate(pairs, 1):
                # creator: 없으면 생성(임시 nickname=seed_key), 있으면 id 만 회수
                cur.execute(SQL_UPSERT_CREATOR, (seed_key, seed_key))
                creator_id = cur.lastrowid
                if not creator_id:
                    # LAST_INSERT_ID 트릭이 있어도 0이 나오는 경우 대비 폴백
                    cur.execute("SELECT creator_id FROM creators WHERE seed_key=%s",
                                (seed_key,))
                    creator_id = cur.fetchone()[0]

                if seed_key in before:
                    existing_creator += 1
                else:
                    created_creator += 1
                    before.add(seed_key)   # 같은 키가 또 나와도 중복 집계 안 되게

                # is_primary: 한 creator의 첫 인스타 채널만 1.
                # 한 사람이 인스타 계정을 여러 개 가진 경우
                # '대표 계정'을 하나만 지정하기 위한 플래그.
                is_primary = 0 if creator_id in primary_done else 1
                primary_done.add(creator_id)

                cur.execute(SQL_UPSERT_CHANNEL, (
                    creator_id, raw, normalize_url(username), is_primary,
                ))
                ch_upserted += 1

                # 500개마다 중간 커밋.
                # 수만 건을 한 트랜잭션에 담으면 undo log가 커지고,
                # 중간에 죽으면 전부 롤백된다. 나눠서 커밋하면
                # 앞부분은 살아남고 재실행 시 ON DUPLICATE로 넘어간다.
                if i % 500 == 0:
                    conn.commit()
                    print(f"  ... {i}/{len(pairs)}")
        conn.commit()
    except Exception:
        conn.rollback()   # 마지막 커밋 이후 부분만 되돌린다
        raise
    finally:
        conn.close()

    print(f"creator 신규 생성 : {created_creator}")
    print(f"creator 기존 사용 : {existing_creator}")
    print(f"channels upsert   : {ch_upserted}")
    print("=" * 62)
    # 다음에 뭘 해야 하는지 알려준다. 파이프라인이 5단계라
    # 순서를 기억하지 않아도 되게.
    print("다음: python -m instagram.steps.import_l1  (L1 크롤 결과를 채널에 반영)")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", default=None, help="엑셀 경로 (기본: reader.DEFAULT_EXCEL)")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    main(args.file, args.dry_run)