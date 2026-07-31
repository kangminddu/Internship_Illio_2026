# -*- coding: utf-8 -*-
"""
reader.py (개선판)

youtube/SNS_정보.xlsx 의 Sheet1 에서 Instagram 컬럼(C)만 읽어
username 리스트를 만든다. DB와 무관하며 아래 형태를 반환한다.

[
    {
        "key": "G_35",
        "username": "fancim_review",
        "url": "https://www.instagram.com/fancim_review/",
        "raw_url": "<원본 셀 값>"
    },
    ...
]

컬럼: A=key, B=TikTok, C=Instagram, D=YouTube

★ "DB와 무관하다"가 이 파일의 설계 포인트다.
------
유튜브 seed.py, 틱톡 seed.py는 엑셀을 읽어 '바로 DB에 넣는다'.
이 파일은 읽어서 '리스트를 반환할 뿐'이다.

  장점 : DB 없이 테스트할 수 있다. python -m instagram.reader로
         엑셀 파싱만 검증 가능하고, 결과를 눈으로 확인한 뒤 적재한다.
  구조 : reader(읽기) → import_l1(적재) 로 책임이 나뉜다.

세 플랫폼이 같은 엑셀의 다른 열을 읽는다:
    A=key(seed_key)  B=TikTok  C=Instagram  D=YouTube
key가 세 플랫폼을 잇는 연결고리라, 한 크리에이터의 여러 채널이
같은 creator_id로 묶인다.
"""

import re
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook


# =========================================================
# 설정
# =========================================================
# 엑셀 기본 경로: reader.py 위치 기준으로 계산 (실행 위치와 무관하게 동작)
#   instagram/reader.py  ->  ../youtube/SNS_정보.xlsx
#
# ⚠️ 주석은 ../youtube/를 가리키는데 코드는 parent.parent / "SNS_정보.xlsx"다.
#    즉 프로젝트 루트의 SNS_정보.xlsx를 본다. 주석과 코드가 어긋난다.
#    (엑셀 위치가 옮겨졌는데 주석만 안 고친 것으로 보임)
DEFAULT_EXCEL = Path(__file__).resolve().parent.parent / "SNS_정보.xlsx"

# username 으로 오면 안 되는 예약 경로 (실제 프로필이 아님)
#
# ★ 화이트리스트가 아니라 블랙리스트다.
#   인스타 URL은 /{username} 형태라 첫 세그먼트가 '아무 문자열'이다.
#   화이트리스트로는 걸러낼 수 없어, 프로필이 아닌 것만 명시하고
#   나머지를 username으로 본다.
#   (유튜브 youtube_url_filter.py도 같은 방식)
#
#   목록이 긴 이유: 시드 엑셀에 /p/(게시물), /reel/(릴스),
#   /explore/ 같은 링크가 실제로 섞여 들어온다.
INVALID_PATHS = {
    "", "accounts", "explore", "reels", "reel", "stories", "story",
    "p", "tv", "channel", "redirect", "invites", "invite",
    "share", "s", "_u", "web", "direct", "directory",
    "about", "legal", "privacy", "terms", "developer", "developers",
    "api", "oauth", "help", "press", "business", "creators",
}

# Instagram username 규칙: 영문/숫자/밑줄/마침표, 1~30자
#
# 유튜브와 대조되는 부분. 유튜브 핸들은 한글·일본어가 들어가서
# 정규식으로 제한할 수 없었지만, 인스타는 ASCII만 허용해서 검증 가능하다.
USERNAME_RE = re.compile(r"^[A-Za-z0-9._]{1,30}$")

# True면 위 규칙에 안 맞는 username을 버린다.
# 기존(관대한) 동작을 그대로 쓰고 싶으면 False 로 두세요.
#
# ★ 스위치로 만들어둔 이유: 규칙을 강화했을 때 기존에 잘 되던 계정이
#   버려질 수 있어, 비교 테스트할 수 있게 남겼다.
VALIDATE_USERNAME = True


# =========================================================
# username 추출
# =========================================================
def extract_username(url, validate=None):
    """
    Instagram URL -> username

    https://www.instagram.com/nike/         -> nike
    https://www.instagram.com/nike/?hl=ko   -> nike
    https://instagram.com/Nike              -> nike   (소문자 정규화)
    https://www.instagram.com/p/XXptr/      -> ""      (게시글 링크)
    "우리 인스타 놀러오세요"                 -> ""      (URL 아님)

    실패 시 빈 문자열을 반환한다.
    (유튜브 url_filter는 (None, 사유) 튜플로 '왜 실패했는지'까지 주는데,
     여기는 그냥 ""다. 아래 load_instagram_urls가 stats로 대략 집계한다)
    """
    if validate is None:
        validate = VALIDATE_USERNAME

    if not url:
        return ""

    url = str(url).strip()

    # URL이 아닌 공유 문구 제거
    # 엑셀에 "우리 인스타 놀러오세요 @nike" 같은 값이 들어온다.
    if "instagram.com" not in url.lower():
        return ""

    # 스킴 보정
    # "instagram.com/nike"처럼 스킴 없이 적힌 값이 흔하다.
    # urlparse는 스킴이 없으면 전체를 path로 인식해 netloc이 빈다.
    if not url.startswith(("http://", "https://")):
        url = "https://" + url

    parsed = urlparse(url)

    # instagram 도메인만 허용
    # 위에서 문자열 포함 검사를 했지만, "evil.com/instagram.com" 같은
    # 경우를 걸러내려면 파싱 후 netloc으로 다시 봐야 한다.
    if "instagram.com" not in parsed.netloc.lower():
        return ""

    path = parsed.path.strip("/")
    if not path:
        return ""       # instagram.com 루트만 있는 경우

    # 첫 세그먼트만 취한다. /nike/tagged/ → nike
    # 소문자로 정규화: 인스타 username은 대소문자를 구분하지 않는데
    # 엑셀에는 "Nike"처럼 적힌 경우가 있다. 정규화 안 하면
    # 같은 계정이 중복 수집된다.
    username = path.split("/")[0].strip().lower()

    # 예약 경로 제거
    if username in INVALID_PATHS:
        return ""

    # 형식 검증 (선택)
    if validate and not USERNAME_RE.match(username):
        return ""

    return username


# =========================================================
# 엑셀 로드
# =========================================================
def load_instagram_urls(excel_path=None, sheet_name="Sheet1", verbose=True):
    """
    Sheet1의 Instagram 컬럼(C)만 읽어 중복 없는 username 리스트를 반환.
    """
    excel_path = Path(excel_path) if excel_path else DEFAULT_EXCEL

    # 파일이 없으면 여기서 멈춘다. 없는 채로 진행하면
    # "대상 0개"만 뜨고 원인을 모른다.
    if not excel_path.exists():
        raise FileNotFoundError(f"엑셀을 찾을 수 없습니다: {excel_path}")

    wb = load_workbook(excel_path, data_only=True)
    # data_only=True: 수식 셀에서 '수식'이 아니라 '계산된 값'을 읽는다.
    # 시드 엑셀에 VLOOKUP 같은 게 있으면 이게 없으면 "=VLOOKUP(...)"이 나온다.

    if sheet_name not in wb.sheetnames:
        wb.close()
        # 존재하는 시트 이름을 함께 알려준다. 시트명이 바뀌었을 때
        # 뭘로 고쳐야 할지 바로 알 수 있다.
        raise ValueError(f"'{sheet_name}' 시트가 없습니다. 존재하는 시트: {wb.sheetnames}")
    ws = wb[sheet_name]

    result = []
    seen = set()
    # ★ 사유별 집계. 이게 곧 시드 엑셀의 품질 리포트가 된다.
    #   "1,881행 중 인스타 URL이 있는 건 몇 개인가"를 숫자로 보여준다.
    stats = {
        "total": 0,
        "empty_or_null": 0,
        "no_username": 0,
        "duplicate": 0,
        "collected": 0,
    }

    for row in ws.iter_rows(min_row=2, values_only=True):
        # min_row=2: 헤더 건너뛰기.
        # ⚠️ 컬럼 위치를 헤더명이 아니라 인덱스로 하드코딩한다(row[2]).
        #    유튜브 seed.py는 헤더명+별칭 사전으로 찾아서 열 순서가
        #    바뀌어도 견딘다. 이쪽은 C열이 밀리면 조용히 엉뚱한 값을 읽는다.
        stats["total"] += 1

        # C열보다 짧은 행 방어
        # openpyxl은 뒤쪽이 전부 빈 행을 짧게 반환한다. 그대로 row[2]를
        # 읽으면 IndexError가 난다.
        if len(row) < 3:
            stats["empty_or_null"] += 1
            continue

        key = row[0]
        instagram = row[2]

        if instagram is None:
            stats["empty_or_null"] += 1
            continue

        raw = str(instagram).strip()
        # 사람이 채운 엑셀에는 빈칸 대신 문자열 "NULL"이 들어온다.
        if not raw or raw.upper() == "NULL":
            stats["empty_or_null"] += 1
            continue

        username = extract_username(raw)
        if not username:
            stats["no_username"] += 1
            continue

        # 같은 계정은 한 번만 (소문자 기준)
        # 여러 크리에이터가 같은 계정을 적어둔 경우(소속사 공식 계정 등)를
        # 걸러낸다. 첫 번째 것만 쓴다.
        if username in seen:
            stats["duplicate"] += 1
            continue

        seen.add(username)
        result.append(
            {
                "key": key,          # seed_key. creators와 연결하는 키
                "username": username,
                "url": f"https://www.instagram.com/{username}/",  # 정규화된 URL
                "raw_url": raw,      # 원본 보존. 나중에 "왜 이렇게 파싱됐지?"를
                                     # 추적할 때 필요하다.
            }
        )
        stats["collected"] += 1

    wb.close()

    if verbose:
        # 한 줄 요약. 엑셀 품질을 바로 확인할 수 있다.
        # (예: no_username이 많으면 다른 플랫폼 주소가 섞였다는 뜻)
        print(
            "[reader] 총 {total}행 | 빈값/NULL {empty_or_null} | "
            "username없음 {no_username} | 중복 {duplicate} | "
            "최종수집 {collected}".format(**stats)
        )

    return result


# l1.py 호환용 별칭 (l1.py의 load_rows에서 get_instagram_rows()로 호출)
#
# 함수 이름을 바꾸면서 호출부를 안 고치려고 별칭을 남겼다.
# 리팩터링 중간 상태. 정리 대상이다.
get_instagram_rows = load_instagram_urls


if __name__ == "__main__":
    # 단독 실행하면 파싱 결과를 확인할 수 있다.
    #   python -m instagram.reader
    # DB를 안 건드리므로 안전하게 몇 번이든 돌려볼 수 있다.
    # 새 엑셀을 받았을 때 이걸 먼저 돌려서 몇 개가 잡히는지 본다.
    rows = load_instagram_urls()
    print(f"총 {len(rows)}개")
    for row in rows[:10]:
        print(row)