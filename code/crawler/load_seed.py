import argparse
import os
import re
import pymysql
from openpyxl import load_workbook

from config import DB



def normalize_url(url):
    if not url:
        return None
    url = str(url).strip()
    url = re.sub(r"\?.*$", "", url)
    for suffix in (
        "/featured", "/videos", "/about", "/discussion",
        "/community", "/playlists", "/streams", "/shorts",
    ):
        if url.endswith(suffix):
            url = url[:-len(suffix)]
    return url.rstrip("/")


def classify_url(url):
    if not url:
        return "unresolved", None, None
    url = normalize_url(url)
    m = re.search(r"(UC[\w-]{22})", url)
    if m:
        return "resolved", url, m.group(1)
    if "/@" in url:
        return "handle_only", url, None
    if "/c/" in url:
        return "custom_only", url, None
    if "/user/" in url:
        return "user_legacy", url, None
    return "unresolved", url, None


def main(xlsx_path):
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["Sheet2"]
    rows = list(ws.iter_rows(values_only=True))

    headers = rows[0]
    rows = rows[1:]

    idx_name = headers.index("상품명 2")
    idx_agency = headers.index("소속")
    idx_url = headers.index("유튜브")
    print(f"Sheet2 데이터 {len(rows)}행\n")

    conn = pymysql.connect(**DB, autocommit=True)
    inserted = 0      # channels까지 만든 수
    creator_only = 0  # creators만 만든 수 (URL 없음)

    with conn.cursor() as cur:
        for i, row in enumerate(rows, 1):
            name = row[idx_name]
            agency = row[idx_agency]
            url = row[idx_url]

            if not name:
                continue

            name = str(name).strip()
            agency = (
                str(agency).strip()
                if agency and str(agency).strip() != "NULL"
                else None
            )
            seed_key = f"S2_{i}"

            # creators는 URL 유무와 상관없이 항상 넣음 (이름·소속 보존)
            cur.execute("""
                INSERT INTO creators (seed_key, nickname, agency)
                VALUES (%s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    nickname=VALUES(nickname),
                    agency=VALUES(agency)
            """, (seed_key, name, agency))

            # URL 없으면 여기서 끝 (channels 안 만듦)
            if url is None or str(url).strip() in ("", "NULL"):
                creator_only += 1
                print(f"CREATOR만  {name:20s} agency={agency or '-':10s} (URL 없음)")
                continue

            # URL 있으면 channels도 생성
            cur.execute(
                "SELECT creator_id FROM creators WHERE seed_key=%s",
                (seed_key,)
            )
            creator_id = cur.fetchone()[0]

            status, url_norm, uc = classify_url(url)

            cur.execute("""
                INSERT INTO channels
                (
                    creator_id,
                    platform,
                    channel_url_raw,
                    channel_url_normalized,
                    channel_id_status,
                    external_channel_id,
                    is_primary
                )
                VALUES
                (
                    %s,'youtube',%s,%s,%s,%s,1
                )
                ON DUPLICATE KEY UPDATE
                    creator_id = VALUES(creator_id),
                    channel_url_normalized = VALUES(channel_url_normalized),
                    channel_id_status = VALUES(channel_id_status),
                    external_channel_id = VALUES(external_channel_id),
                    is_primary = VALUES(is_primary)
            """, (
                creator_id,
                str(url).strip(),
                url_norm,
                status,
                uc
            ))

            inserted += 1
            print(
                f"{name:20s} "
                f"[{status:12s}] "
                f"agency={agency or '-':10s} "
                f"uc={uc or '-'}"
            )

    conn.close()
    print("\n======================")
    print(f"channels 생성 : {inserted}")
    print(f"creators만    : {creator_only}")
    print(f"creators 총계 : {inserted + creator_only}")
    print("======================")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="YouTube Seed Loader"
    )

    parser.add_argument(
        "--file",
        required=True,
        help="Seed Excel (.xlsx)"
    )

    args = parser.parse_args()

    if not os.path.exists(args.file):
        raise FileNotFoundError(args.file)

    main(args.file)