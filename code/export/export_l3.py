import os
import warnings
warnings.filterwarnings("ignore", category=UserWarning)
import pandas as pd
import pymysql

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config import DB, EXPORT_DIR

sql = """
SELECT

ROW_NUMBER() OVER(
ORDER BY
cr.nickname,
ct.published_at DESC,
c.published_at DESC
) AS '번호',

cr.nickname AS '크리에이터',

ct.external_id AS '영상ID',

c.external_comment_id AS '댓글ID',

f.external_author_id AS '댓글작성자UC',

c.author_display_name AS '댓글작성자',

DATE_FORMAT(
c.published_at,
'%Y-%m-%d %H:%i'
) AS '댓글작성일',

c.like_count AS '댓글좋아요',

CHAR_LENGTH(c.comment_text) AS '댓글길이',

c.comment_text AS '댓글내용'

FROM comments c

JOIN fans f
ON c.fan_id=f.fan_id

JOIN contents ct
ON c.content_id=ct.content_id

JOIN channels ch
ON ct.channel_id=ch.channel_id

JOIN creators cr
ON ch.creator_id=cr.creator_id

ORDER BY
cr.nickname,
ct.published_at DESC,
c.published_at DESC;
"""

# ------------------------------------------------
# DB -> DataFrame
# ------------------------------------------------

conn = pymysql.connect(**DB)

df = pd.read_sql(sql, conn)

conn.close()

filename = os.path.join(EXPORT_DIR, "L3_댓글정보_포카챠.xlsx")

df.to_excel(
    filename,
    index=False,
    sheet_name="L3"
)

# ------------------------------------------------
# Excel Styling
# ------------------------------------------------

wb = load_workbook(filename)
ws = wb["L3"]

header_fill = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
)

header_font = Font(
    bold=True,
    color="FFFFFF"
)

center = Alignment(
    horizontal="center",
    vertical="center"
)

thin = Side(
    border_style="thin",
    color="D9D9D9"
)

border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)

# -------------------------------
# Header
# -------------------------------

for cell in ws[1]:

    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

# -------------------------------
# Body
# -------------------------------

for row in ws.iter_rows(min_row=2):

    for cell in row:

        cell.border = border

        if cell.column in [1,7,8,9]:
            cell.alignment = center

# -------------------------------
# Number Format
# -------------------------------

for col in ["H","I"]:

    for cell in ws[col][1:]:

        cell.number_format = "#,##0"

# -------------------------------
# Column Width
# -------------------------------

widths = {

    "A":8,      # 번호
    "B":20,     # 크리에이터
    "C":18,     # 영상ID
    "D":28,     # 댓글ID
    "E":30,     # 댓글작성자UC
    "F":20,     # 댓글작성자
    "G":20,     # 댓글작성일
    "H":12,     # 댓글좋아요
    "I":12,     # 댓글길이
    "J":80      # 댓글내용

}

for col,width in widths.items():

    ws.column_dimensions[col].width = width

# -------------------------------
# Freeze
# -------------------------------

ws.freeze_panes = "A2"

# -------------------------------
# Filter
# -------------------------------

ws.auto_filter.ref = ws.dimensions

wb.save(filename)

print(f"완료 : {filename}")
print(f"댓글 {len(df):,}개")