import os
import warnings
warnings.filterwarnings("ignore", category=UserWarning)
import pandas as pd
import pymysql

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import DB, EXPORT_DIR

sql = """
SELECT
    ROW_NUMBER() OVER (
        ORDER BY cs.follower_count DESC
    ) AS '번호',

    cr.nickname AS '크리에이터',

    COALESCE(cr.agency,'-') AS '소속',

    CASE
        WHEN ch.channel_activity_status='active'
            THEN '활동중'
        WHEN ch.channel_activity_status='low_active'
            THEN '저활동'
        WHEN ch.channel_activity_status='inactive'
            THEN '휴면'
        ELSE '-'
    END AS '활동상태',

    cs.follower_count AS '구독자수',

    cs.total_view_count AS '누적조회수',

    cs.total_video_count AS '총영상수',

    DATE_FORMAT(ch.channel_opened_at,'%Y-%m-%d') AS '채널개설일',

    -- ▼ 추가: 크롤 결과
    CASE
        WHEN log.status='success' THEN '성공'
        WHEN log.status='failed'  THEN '실패'
        ELSE '-'
    END AS '수집결과',

    COALESCE(log.error_type,'-')  AS '실패사유',

    COALESCE(log.http_status,'-') AS 'HTTP코드',

    COALESCE(ch.channel_url_normalized, ch.channel_url_raw) AS 'URL'

FROM creators cr

JOIN channels ch
ON cr.creator_id = ch.creator_id

LEFT JOIN (
    SELECT cs1.*
    FROM channel_snapshots cs1
    JOIN (
        SELECT
            channel_id,
            MAX(captured_at) latest
        FROM channel_snapshots
        GROUP BY channel_id
    ) t
    ON cs1.channel_id=t.channel_id
    AND cs1.captured_at=t.latest
) cs
ON ch.channel_id=cs.channel_id

-- ▼ 추가: 각 채널의 최신 L1 크롤 기록
LEFT JOIN (
    SELECT c.channel_id, c.status, c.error_type, c.http_status
    FROM crawl_logs c
    JOIN (
        SELECT channel_id, MAX(log_id) AS mx
        FROM crawl_logs
        WHERE layer='L1' AND channel_id IS NOT NULL
        GROUP BY channel_id
    ) m ON c.log_id = m.mx
) log
ON log.channel_id = ch.channel_id

ORDER BY cs.follower_count DESC;
"""

# ----------------------------------------
# DB → DataFrame
# ----------------------------------------

conn = pymysql.connect(**DB)

df = pd.read_sql(sql, conn)

conn.close()

filename = os.path.join(EXPORT_DIR, "L1_채널정보_상세.xlsx")

df.to_excel(
    filename,
    index=False,
    sheet_name="L1"
)

# ----------------------------------------
# Excel 꾸미기
# ----------------------------------------

wb = load_workbook(filename)
ws = wb["L1"]

# ===== 스타일 =====

header_fill = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
)

header_font = Font(
    bold=True,
    color="FFFFFF"
)

center = Alignment(
    horizontal="center",
    vertical="center"
)

thin = Side(
    border_style="thin",
    color="D9D9D9"
)

border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)

# ===== 헤더 =====

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

# ===== 전체 셀 =====

for row in ws.iter_rows(min_row=2):

    for cell in row:

        cell.border = border

        if cell.column in [1,5,6,7]:
            cell.alignment = center

# ===== 숫자 서식 =====

for col in ["E","F","G"]:

    for cell in ws[col][1:]:

        cell.number_format = '#,##0'

# ===== 열 너비 자동 =====

for column_cells in ws.columns:

    length = max(
        len(str(cell.value)) if cell.value is not None else 0
        for cell in column_cells
    )

    letter = get_column_letter(column_cells[0].column)

    ws.column_dimensions[letter].width = min(length + 4, 35)

# ===== 첫 행 고정 =====

ws.freeze_panes = "A2"

# ===== 필터 =====

ws.auto_filter.ref = ws.dimensions

wb.save(filename)

print(f"완료 : {filename}")
print(f"총 {len(df)}개 채널")