import os
import warnings
warnings.filterwarnings("ignore", category=UserWarning)
import pandas as pd
import pymysql

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =====================================
# DB
# =====================================

from config import DB, EXPORT_DIR

# =====================================
# SQL
# =====================================

sql = """
SELECT

ROW_NUMBER() OVER(
ORDER BY cr.nickname
) AS 번호,

cr.nickname AS 크리에이터,

COALESCE(cr.agency,'-') AS 소속,

MAX(chs.follower_count) AS follower_count,

COUNT(DISTINCT ct.content_id) AS video_count,

AVG(cs.view_count) AS avg_view,

AVG(cs.like_count) AS avg_like,

AVG(cs.comment_count) AS avg_comment,

m.commenter_overlap_rate,

m.regular_commenter_count,

m.avg_comment_length

FROM creators cr

JOIN channels ch
ON cr.creator_id = ch.creator_id

LEFT JOIN contents ct
ON ch.channel_id = ct.channel_id

LEFT JOIN
(
    SELECT cs1.*

    FROM content_snapshots cs1

    JOIN
    (
        SELECT
            content_id,
            MAX(captured_at) latest

        FROM content_snapshots

        GROUP BY content_id

    ) t

    ON cs1.content_id=t.content_id
    AND cs1.captured_at=t.latest

) cs

ON ct.content_id=cs.content_id


LEFT JOIN
(
    SELECT chs1.*

    FROM channel_snapshots chs1

    JOIN
    (
        SELECT
            channel_id,
            MAX(captured_at) latest

        FROM channel_snapshots

        GROUP BY channel_id

    ) t

    ON chs1.channel_id=t.channel_id
    AND chs1.captured_at=t.latest

) chs

ON ch.channel_id=chs.channel_id


LEFT JOIN channel_metrics m

ON ch.channel_id=m.channel_id

GROUP BY

ch.channel_id

ORDER BY

cr.nickname;
"""

# =====================================
# SQL → DataFrame
# =====================================

conn = pymysql.connect(**DB)

df = pd.read_sql(sql, conn)

conn.close()
print(df.head())
print(df.columns)

# =====================================
# 파생지표 계산 함수
# =====================================

def calc_view_ratio(row):
    """구독자 대비 평균 조회수 비율 (%)"""
    if pd.isna(row["follower_count"]) or row["follower_count"] == 0:
        return "N/A"

    if pd.isna(row["avg_view"]):
        return "N/A"

    return round(
        row["avg_view"] / row["follower_count"] * 100,
        2
    )


def calc_er(row):
    """공개 참여율 (ER)"""
    if pd.isna(row["avg_view"]) or row["avg_view"] == 0:
        return "N/A"

    like = row["avg_like"] if not pd.isna(row["avg_like"]) else 0
    comment = row["avg_comment"] if not pd.isna(row["avg_comment"]) else 0

    return round(
        (like + comment)
        / row["avg_view"]
        * 100,
        2
    )


def calc_upload(row):
    """업로드 빈도 (최근 6개월 = 26주 기준)"""
    if pd.isna(row["video_count"]):
        return "N/A"

    return round(
        row["video_count"] / 26,
        2
    )


def calc_loyalty(row):
    """Loyalty Score"""

    if pd.isna(row["avg_view"]) or row["avg_view"] == 0:
        return "N/A"

    like = row["avg_like"] if not pd.isna(row["avg_like"]) else 0
    comment = row["avg_comment"] if not pd.isna(row["avg_comment"]) else 0

    return round(
        (comment * 10 + like)
        / row["avg_view"],
        4
    )


# =====================================
# 계산
# =====================================

df["구독자대비조회율(%)"] = df.apply(calc_view_ratio, axis=1)

df["공개참여율(ER)"] = df.apply(calc_er, axis=1)

df["업로드빈도(주)"] = df.apply(calc_upload, axis=1)

df["Loyalty Score"] = df.apply(calc_loyalty, axis=1)


# =====================================
# L3 컬럼
# =====================================

df["댓글작성자중복률(%)"] = df["commenter_overlap_rate"]

df["고정댓글러"] = df["regular_commenter_count"]

df["평균댓글길이"] = df["avg_comment_length"]


# =====================================
# L3 없는 채널 → N/A
# =====================================

for col in [
    "댓글작성자중복률(%)",
    "고정댓글러",
    "평균댓글길이"
]:
    df[col] = df[col].apply(lambda x: "N/A" if pd.isna(x) else x)


# =====================================
# 최종 컬럼
# =====================================
print(
    df.loc[
        df["크리에이터"].isin(["깅나리", "디아나", "모이 M0I", "므르크스 므나"]),
        [
            "크리에이터",
            "follower_count",
            "avg_view",
            "avg_like",
            "avg_comment",
            "구독자대비조회율(%)",
            "공개참여율(ER)",
            "업로드빈도(주)",
            "Loyalty Score"
        ]
    ]
)
df = df[
[
    "번호",
    "크리에이터",
    "소속",
    "구독자대비조회율(%)",
    "공개참여율(ER)",
    "업로드빈도(주)",
    "댓글작성자중복률(%)",
    "고정댓글러",
    "평균댓글길이",
    "Loyalty Score"
]
]



# =====================================
# Excel 저장
# =====================================

filename = os.path.join(EXPORT_DIR, "파생지표.xlsx")

df.to_excel(
    filename,
    index=False,
    sheet_name="Metrics"
)

wb = load_workbook(filename)
ws = wb["Metrics"]

# =====================================
# 스타일
# =====================================

header_fill = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
)

header_font = Font(
    bold=True,
    color="FFFFFF"
)

center = Alignment(
    horizontal="center",
    vertical="center"
)

thin = Side(
    border_style="thin",
    color="D9D9D9"
)

border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)

# =====================================
# 헤더 스타일
# =====================================

for cell in ws[1]:

    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

# =====================================
# 본문 스타일
# =====================================

for row in ws.iter_rows(min_row=2):

    for cell in row:

        cell.border = border
        cell.alignment = center

# =====================================
# 숫자 서식
# =====================================

# 구독자대비조회율
for cell in ws["D"][1:]:

    if isinstance(cell.value,(int,float)):
        cell.number_format="0.00"

# ER
for cell in ws["E"][1:]:

    if isinstance(cell.value,(int,float)):
        cell.number_format="0.00"

# 업로드빈도
for cell in ws["F"][1:]:

    if isinstance(cell.value,(int,float)):
        cell.number_format="0.00"

# 댓글중복률
for cell in ws["G"][1:]:

    if isinstance(cell.value,(int,float)):
        cell.number_format="0.00"

# 고정댓글러
for cell in ws["H"][1:]:

    if isinstance(cell.value,(int,float)):
        cell.number_format="#,##0"

# 평균댓글길이
for cell in ws["I"][1:]:

    if isinstance(cell.value,(int,float)):
        cell.number_format="0.0"

# Loyalty
for cell in ws["J"][1:]:

    if isinstance(cell.value,(int,float)):
        cell.number_format="0.0000"
        
# =====================================
# 열 너비
# =====================================

widths = {
    "A": 8,
    "B": 22,
    "C": 18,
    "D": 20,
    "E": 18,
    "F": 18,
    "G": 22,
    "H": 14,
    "I": 16,
    "J": 16
}

for col, width in widths.items():
    ws.column_dimensions[col].width = width


# =====================================
# 첫 행 고정
# =====================================

ws.freeze_panes = "A2"


# =====================================
# 필터
# =====================================

ws.auto_filter.ref = ws.dimensions


# =====================================
# 저장
# =====================================

wb.save(filename)

print("=" * 60)
print(f"완료 : {filename}")
print(f"채널 수 : {len(df)}")
print("=" * 60)

import os
print(os.path.abspath(filename))

check = pd.read_excel(filename)
print(check.loc[check["크리에이터"] == "깅나리"].to_string(index=False))