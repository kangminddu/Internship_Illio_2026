import os
import warnings
warnings.filterwarnings("ignore", category=UserWarning)
import pandas as pd
import pymysql

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config import DB, EXPORT_DIR

sql = """
SELECT

ROW_NUMBER() OVER(
ORDER BY
cr.nickname,
ct.published_at DESC
) AS '번호',

cr.nickname AS '크리에이터',

COALESCE(cr.agency,'-') AS '소속',

DATE_FORMAT(
ct.published_at,
'%Y-%m-%d'
) AS '게시일',

DATEDIFF(
CURDATE(),
ct.published_at
) AS '업로드경과일',

cs.view_count AS '조회수',

cs.like_count AS '좋아요',

cs.comment_count AS '댓글수',

ct.duration_sec AS '길이(초)',

ct.category AS '카테고리',

CONCAT(
'https://www.youtube.com/watch?v=',
ct.external_id
) AS '영상URL'

FROM creators cr

JOIN channels ch
ON cr.creator_id = ch.creator_id

JOIN contents ct
ON ch.channel_id = ct.channel_id

LEFT JOIN (

SELECT cs1.*

FROM content_snapshots cs1

JOIN (

SELECT
content_id,
MAX(captured_at) latest

FROM content_snapshots

GROUP BY content_id

) t

ON cs1.content_id=t.content_id
AND cs1.captured_at=t.latest

) cs

ON ct.content_id=cs.content_id

ORDER BY
cr.nickname,
ct.published_at DESC;
"""

# ----------------------------
# DB -> DataFrame
# ----------------------------

conn = pymysql.connect(**DB)

df = pd.read_sql(sql, conn)

conn.close()

# ----------------------------
# 영상길이 생성 (hh:mm:ss)
# ----------------------------

df["영상길이"] = (
    pd.to_timedelta(df["길이(초)"], unit="s")
    .astype(str)
    .str.replace("0 days ", "", regex=False)
)

# 영상길이를 길이(초) 뒤로 이동

cols = list(df.columns)

cols.remove("영상길이")

idx = cols.index("길이(초)") + 1

cols.insert(idx, "영상길이")

df = df[cols]

filename = os.path.join(EXPORT_DIR, "L2_영상정보_포카챠.xlsx")

df.to_excel(
    filename,
    index=False,
    sheet_name="L2"
)

# =====================================================
# Excel Styling
# =====================================================

wb = load_workbook(filename)
ws = wb["L2"]

header_fill = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
)

header_font = Font(
    bold=True,
    color="FFFFFF"
)

center = Alignment(
    horizontal="center",
    vertical="center"
)

thin = Side(
    border_style="thin",
    color="D9D9D9"
)

border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)

# ----------------------------
# 헤더
# ----------------------------

for cell in ws[1]:

    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

# ----------------------------
# 본문
# ----------------------------

for row in ws.iter_rows(min_row=2):

    for cell in row:

        cell.border = border

        if cell.column in [1,4,5,6,7,8,9,10]:
            cell.alignment = center

# ----------------------------
# 숫자 포맷
# ----------------------------

for col in ["F","G","H","I"]:

    for cell in ws[col][1:]:

        cell.number_format = "#,##0"

# ----------------------------
# 열 너비
# ----------------------------

widths = {

    "A":8,
    "B":20,
    "C":16,
    "D":14,
    "E":14,
    "F":14,
    "G":12,
    "H":12,
    "I":12,
    "J":14,
    "K":20,
    "L":55

}

for col,width in widths.items():

    ws.column_dimensions[col].width = width

# ----------------------------
# 첫행 고정
# ----------------------------

ws.freeze_panes = "A2"

# ----------------------------
# 필터
# ----------------------------

ws.auto_filter.ref = ws.dimensions

wb.save(filename)

print(f"완료 : {filename}")
print(f"영상 {len(df)}개")